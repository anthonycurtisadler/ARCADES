# Simple program for facilitating the creation of book indexes.
# To help determine the key words, establish page numbers,
# and format the final index. 

try:    
    import pdfplumber
except:
    print('MUST INSTALL PDFPLUMBER')

import os, string, simple_parser_mod as parser, pickle
import copy
import nltk
from nltk import word_tokenize
from openpyxl import load_workbook
import itertools

from indexdisplay import IndexDisplay 



from lexical import English_frequent_words
from spellchecker import SpellChecker

from numbertools  import format_range, rom_to_int,\
     int_to_roman,de_range,abbreviate_range,convert_range
from scroll import Select
from globalconstants import YESTERMS 

from indexformat import FormatIndex
from stack import Stack

from indexconfiguration import LEFT_ITALIC, RIGHT_ITALIC, \
     DISPLAY_COLOR, NAME_WORDS, SUBHEAD_WORDS, BREAK_PHRASE, SEARCH_HELP,\
     DIVIDER, SMALL_WORDS, ROWS_IN_PAGE

from indexer_read_me import READ_ME



first_romans = [int_to_roman(x) for x in range(1,300) if x]

BREAK_PHRASE_LEN = len(BREAK_PHRASE)
                    

max_title_length = 15

directoryname = os.getcwd()
folder = os.altsep+'pdfs'+os.altsep
index_folder = os.altsep+'indexer'+os.altsep
speller = SpellChecker()
os.system(DISPLAY_COLOR)



YESTERMS += [' ','Y']
string.punctuation += '—'+'“'+'”'+'‘'+'’'+''
string.punctuation = string.punctuation.replace('-','')
string.punctuation = string.punctuation.replace('’','')
string.punctuation = string.punctuation.replace("'",'')
                        


from colorama import Fore, Back, Style

def box (x):

    
    

    if isinstance(x,str):
        box_list = x.split('\n')
    else:
        box_list = list(x)
    max_len = max([len(l) for l in box_list])

    for index in range(len(box_list)):

        box_list[index] = BOX_CHAR['v']+box_list[index]+(max_len-len(box_list[index]))*' '+BOX_CHAR['v']
        

    top = BOX_CHAR['lu']+BOX_CHAR['h']*max_len+BOX_CHAR['ru']
    
    bottom = BOX_CHAR['ll']+BOX_CHAR['h']*max_len+BOX_CHAR['rl']
    
    return '\n'.join([top]+box_list+[bottom])

    
    
        


def line_form (x,length=70,breaker=';'):

                """Formats  a section of text for each line is no more than legnth"""
                

                return_lines = []
                last_line = ''
                last_line_length = 0
                break_len = len(breaker)
                

                for l in x.split(breaker):
                    if last_line_length + len(l) < length:
                        last_line+=l+breaker
                        last_line_length += len(l)+break_len
                        
                    else:
                        return_lines.append(last_line)
                        last_line = l+breaker
                        last_line_length = len(l)+break_len
                return_lines.append(last_line)
                return '\n'.join(return_lines)
            
def page_sort_function(x):

    """For order roman and italic pages"""

    if not isinstance(x,str):
        return x

    if x.isnumeric():
        return int(x)
    if not set(x)-{'i','x','v','l','c','d'}:
        return -100+rom_to_int(x)
    else:
        return x
    
def format_result(search_term='',page_set=None,page_dict=None,rejected=None):

    """For formatting the results of an index and its subterms for cutting and
        pasting into excel so that it can be interpreted properly"""

    as_line = ''
    if page_set:
        print('{'+format_range(page_set)+'}')
        as_line += '{'+format_range(page_set)+'}'+'\t'
    else:
        as_line += '{%}\t'
    if page_dict:
        classified_list = []
        for classifier in page_dict:
            if page_dict[classifier]:
                classified_list.append(classifier.split('/')[0]+'{'+format_range(page_dict[classifier])+'}')
        print(';'.join(classified_list))
        as_line += ';'.join(classified_list)
    if rejected:
        print('REJECTED ','{-'+','.join(rejected)+'}')
    return as_line

def numeric_input(prompt='',lower=None,upper=None,alert=True):

    """For inputting a numeric value"""

    while True:
        x = input(prompt)
        if x.isnumeric():
            if (lower and int(x)<lower)  or (upper and int(x)>upper):
                print('VALUE OUT OF BOUNDS! MUST BE BETWEEN '+str(lower)+' AND '+str(upper)+'!')
            else:
                break
        elif alert:
            print('VALUE MUST BE AN INTEGER!')
    return x

def type_input(prompt='',must_be=['Y','N',' '],truncate=True,upper=True,return_empty=True, alert=True):

    """For inputting limited to a selection of single letter values"""
    
    while True:
        x = input(prompt)
        if x and truncate:
            x = x[0]
        if x and upper:
            x = x.upper()
            
        if x in must_be or (not x and return_empty):
            return x
        elif alert:
            print('VALUE MUST BE '+','.join(must_be).replace(' ','BLANK')+',RETURN'*return_empty)
        

def yes_no_input(prompt=''):

    """For yes or no inputs"""

    return type_input(prompt) in [' ','Y']


def get_if(x,left=None,right=None,get_all=False):

    """Extracts the segment from a text surrounded by left and right.
    Returns extract, text"""

    if not get_all:

        if left and right:
            if left in x and right in x:
                return x.split(left)[1].split(right)[0].strip(),\
                       x.split(left)[0]+right.join(x.split(right)[1:])
            
            return '',x.strip()
    all_found = []
    if left and right:
        counter = 0
        while  counter < 10 and left in x and right in x:
            counter += 1 
            found, x = get_if(x,left,right,get_all=False)
            all_found.append(found)
        return list(set(all_found)), x
    return [],''
            
            

def purge_small (phrase,purge_list,italics=True):

    """Purge all words from the purge list.
        and remove italic markers"""
    
    if italics:
        phrase = phrase.replace(LEFT_ITALIC,'').replace(RIGHT_ITALIC,'')
        
    
    surround = lambda a:' '+a+' '
    phrase = surround(phrase)
    for x in purge_list:
        if surround(x) in phrase:
            phrase = phrase.replace(surround(x),' ')
    return phrase.strip() 

def some_letters (x):

    """Returns TRUE if a phrase has some characters that are
    not whitespace or punctuation."""

    letters = set(x)
    for char in letters:
        if char not in string.whitespace+string.punctuation:
            return True
    return False

def strip_punctuation (phrase,left=True,right=True):

    """Strips punctuation to right or left of phrase"""

    if left:
        while phrase:
            if phrase[0] in string.punctuation:
                phrase = phrase[1:]
            else:
                break
    if right:
        while phrase:
            if phrase[-1] in string.punctuation:
                phrase = phrase[:-1]
            else:
                break
    return phrase 


def extract_page_number (text,from_top=False):

    """Determines the first numeric string in either
    the top or the bottom of the text"""

    # to test whether the text has not blank and non
    # punctuation characters
    if not text:
        return ''
    text = text.rstrip()
    text = text.replace('\n',' ')
    ghost_text = text
    for x in string.punctuation+string.whitespace:
        ghost_text = ghost_text.replace(x,'')
    if not ghost_text:
        return ''
    
        
    if ' ' not in text:
        return ''
    numeric_found = False
    for x in '0123456789ivx':
        if x in text:
            numeric_found = True
    if not numeric_found:
        return ''

    #Testing for the simplest case,
    #Namely that the page number is
    #immediately at top or bottom
    #of the page
    
    if not from_top:
    
        page_no = text.split(' ')[-1]
    else:
        page_no = text.split(' ')[0]

    if page_no.isnumeric() or page_no in first_romans:
        return page_no

    if not from_top:
        text = ''.join(reversed(text))

    #In case the page no. is not
    #immediately at the top or bottom
    #of the page 
    numeric_found = False
    page_no = ''
    for char in text:
        if char.isnumeric():
            page_no += char
            if not numeric_found:
                numeric_found = True
        else:
            if numeric_found:
                break
    if not from_top:
        page_no = ''.join(reversed(page_no))
    
    
    return page_no

def is_date (x):

    """Returns TRUE if a string
    is formatted like a date.
    """

    x = x.replace('-','').replace('–','').strip()
    return x.isnumeric()

def un_date (x):

    """Removes the date/date range from the
    end of a book title"""

    while ',' in x and  last_is_date(x):
        x = ','.join(x.split(',')[0:-1])
    return x

def last_is_date (x):

    """True if a string ends with a date/date-range"""
    
    if not ',' in x:
        return True
    return is_date(x.split(',')[-1])

def sort_all_pages (all_pages):
    
        roman_pages = [int_to_roman(x)
                               for x in sorted([rom_to_int(x) for x in all_pages
                                                if not x.isnumeric()])]
        arabic_pages = [str(x) for x in sorted([int(x) for x in all_pages
                                                        if x.isnumeric()])]
        all_pages = roman_pages + arabic_pages
        return all_pages


for x in list(SMALL_WORDS):
    SMALL_WORDS.append(x[0].upper()+x[1:])


class Reader:

    """ Interprets a PDF into text divided into pages, indexed words,
    sentences, quoted phrases, parenthesized phrases, bracketed phrases.
    The READER class is instantiated as a TEXT object, which contains all the
    basic information about the TEXT"""
    
    

    def __init__ (self,database_object=None,in_italics=True,in_quotes=True,in_parens=True,in_caps=True,all_sentences=True,above=None,below=None):

        self.database_object = database_object
        self.in_italics = in_italics
        self.in_quotes = in_quotes
        self.all_sentences = all_sentences
        self.in_parens = in_parens
        self.in_caps = in_caps
        self.above = above
        self.below = below
        

        self.page_record = [] #list of pages found in PDF
        self.sentence_record = [] #List of sentences found
                                  #consisting of a tupple:
                                  #(starting position,
                                  #(finishing position,
                                  #(sentence count)
        
        self.page_dict = {}       #keys = page #
                                  #values = (starting position,
                                  #          ending position)
        self.sentences = set()    #Set of sentence strings
        self.brackets = set()     #Set of bracketed phrases       
        self.sentence_dict = {}
        
        self.pages = {}           #keys = page #
                                  #value = page text
        self.italicized_chars = set() #keeps track of positions in
                                      #text of italicized characters
        self.italicized_phrases = set()
        self.parenthetical_phrases = set()
        self.bracketed_phrases = set()
        self.capitalized_phrases = set() 
        self.quoted_phrases = set()
        self.all_text = ''       #The entire text as a single object      
        self.words = {}          #keys = words / values = page#
        self.words_in_sentences = {} #keys = words / values = sentence#
        self.nouns = set()           #identifies all nouns
        self.histio = {}             #keeps track of word frequency
        self.italic_dict = {}      #keeps track of pages in which italic expressions are found
        self.title_dict = {}      #keeps track of pages in which titles are found 
        self.capital_histio = {}  #keeps track of capitalized expressions 
        self.first_words = set()  #first words in sentences
        self.dehyphenated = {} #Keeps track of dehyphenated expressions
        self.order_dictionary = {} #Keeps track of first appearance of phrases and words
        self.title_translation_dictionary = {} # keeps track of likely translations of titles
        self.author_dict = {} # keeps track of likely authors

        self.last_working_title = ''
        self.working_text = ''
        
        
        

        ##This data structure is very messy; could be simplified, cleaned up;
        ##too much redundancy.

    def is_greater_than (x_page,y_page):

        if not x_page.isnumeric():
            if y_page.isnumeric():
                return False
            elif rom_to_int(x_page) <= rom_to_int(y_page):
                return False
            else:
                return True
        elif not y_page.isnumeric():
            return True 
        else: return not int(x_page) <= int(y_page)


        
        
    def get_sentences(self,sentence_list=None,text_title=None):

        """Returns pages in which a sentence can be found"""


        
        return_pages = set()
        for s in sentence_list:
            if not self.database_object:
                if s in self.sentence_dict:
                    return_pages.update(set(self.sentence_dict[s][2].split(',')))
            elif text_title:
                self.database_object.cursor.execute("SELECT * FROM pages_for_sentences WHERE book=? AND sentence=",(text_title,s,))
                x = self.database_object.cursor.fetchone()[0]
                print('PAGES FOR SENTENCES',x)
                return_pages.update(x)
                
        return return_pages

        
    def load (self,filename='test.pdf',text_title=''):

        """Reads a PDF file and analyzes it,
        dividing it into page, and extracting
        italicized, bracketed, quoted, etc.
        phrases, and well as creating a general histogram,
        and dividing it into sentences."""

        def get_parameters ():

            """Query user for the parameters for intepreting"""
            from_top = yes_no_input(prompt='Is the page number on top of the page?') 
            add_italic_caps = yes_no_input(prompt='Include italicized capital expressions?')

            return from_top, add_italic_caps

        def get_all_fonts_from_page (page_data):

            """Retrieves unique fonts from a single page"""
            
            fonts = set()
            for char in page_data.chars:
                if char['fontname'] not in fonts:
                    fonts.add(char['fontname'])
            return fonts

        def get_all_fonts_from_text (text,from_this=0,to_that=1000):

            """To retrieve all unique fonts from all texts"""
            
            all_fonts = set()
            query = True
            for counter, page in enumerate(text.pages[from_this:to_that]):
                print(DIVIDER+counter)
                found = get_all_fonts_from_page(page)
                print(found)
                if query:
                    inp = input('RETURN to continue, SPACE+RETURN to stop querying! ')
                    if inp:
                        query = False
                all_fonts.update(found)
            return all_fonts

        def determine_exclude_set (font_set):

            """To query which fonts are to be excluded"""
            
            return_set = set()
            print('ALL FONTS: ',','.join(font_set))
            for f in sorted(font_set):
                if yes_no_input('EXCLUDE'+f+'? '):
                    return_set.add(f)
            return return_set

        

     
        position_at = 0
        sentence_starts_at = 0
        found_left_paren = False
        found_left_bracket = False
        found_left_quote = False
        paren_string = ''
        bracket_string = ''
        quoted_string = ''
        sentence_string = []
        sentence_count = 0
        word_starting = False
        capitalized_words = []
        last_capitalized = False
        word_position_in_sentence = 0
        italic_pages = []
        last_not_alpha = False
        if not filename.endswith('.pdf'):
            filename+= '.pdf'
        excluded_fonts = set()
        sentence_pages = set()
        last_title = None
        from_last_title = 0
        from_last_cap_phrase = 0
        char_set = set()
        last_char = ''
        last_cap_phrase = ''
        last_alpha_x0 = 0
        last_alpha_y0 = 0
        

            
        
        word = ''
        text_to_exclude = []

        if self.database_object:
            self.database_object.add_book(text_title)
        
        
        first_page = True
        last_found_page = None
        if not self.database_object:
            from_top, add_italic_caps = get_parameters()
        else:
            from_top, add_italic_caps = True, True
        def get_limits (pdf,from_this,to_that):

            """To make sure page limit is within bounds
            """
            
            if from_this >= len(pdf.pages):
                from_this = len(pdf.pages)-1
            if to_that >= len(pdf.pages):
                to_that = len(pdf.pages)-1
                            
            return from_this, to_that
        
        def proper_page (x,to=False):

            """Returns an integer for a numeric string,
            otherwise 0 or high upperbound"""

            if x.isnumeric():
                x = int(x)
            else:
                if to:
                    x = 10000000000000 
                else:
                    x = 0
            return x

        def de_hyphen (x):

            if '-' not in x:
                return x
            if x.count('-') == 1 and x.split('-')[1].islower():
                dh = x.replace('-','')
                if  not speller.unknown({dh}):
                    x=dh
                return x
            else:
                
                return de_hyphen(x.split('-')[0])+'-'+de_hyphen('-'.join(x.split('-')[1:]))

        def determine_page_numeration (pdf,above=True,below=False):

            page_dict = {}
            def get_number (x,up_to=50):
                if not isinstance(x,str):
                    return 0
                

                number = ''
                starting = False
                counter = 0
                for c in x:
                    if counter > up_to:
                        break
                    if not starting and c.isnumeric():
                        starting = True
                        number+=c
                    elif starting and c.isnumeric():
                        number+=c
                    elif starting and not c.isnumeric():
                        return int(number)
                    counter += 1
                    
                return 0

            def get_possible_pages(pdf):

                page_list = pdf.pages
                
                page_dict = {}
                for counter,page in  enumerate(page_list):
                    if counter%10==0:
                        print (counter)
                    if not page is None:
                        text = page.extract_text()
                        lower_number,upper_number = 0,0
                        if text:
                            if above:
                                upper_number = get_number (text)
                            if below:
                                threshold = len(text)
                                if threshold>50:
                                    threshold = 50 
                                lower_number = get_number (''.join(reversed(text[-threshold:])))
                                
                                    
                    else:
                        upper_number = 0
                        lower_number = 0

                    page_dict[counter] = (upper_number,lower_number)


                
                return page_dict 

            def get_page_lists(page_dict,index=0):

                
                all_lists = []
                working_list = []
                before_last = None
                last = None
                is_adjacent = lambda x,y: y[0]-x[0] == y[1]-x[1] 
                for counter, p in enumerate(page_dict):
                    if counter%10==0:
                        print (counter)
                    current = (p,page_dict[p][index])
                    
                    if not working_list and last and is_adjacent(last,current):
                       working_list.append((p-1,last[1]))
                       working_list.append((p,current[1]))
                    elif not working_list and before_last and is_adjacent(before_last,current):
                       working_list.append((p-2,before_last[1]))
                       working_list.append((p-1,current[1]-1))
                       working_list.append((p,current[1]))
                    elif working_list and working_list[-1][0]==p-1 and is_adjacent(working_list[-1],current):
                       working_list.append((p,current[1]))
                    elif working_list:
                       all_lists.append(working_list)
                       working_list = []
                    if last:
                        before_last = last
                    last = current
                  
                if working_list:
                    all_lists.append(working_list)
                
                return all_lists
            
            def find_longest_string (x):
                try:

                    return [y for y in x if len(y) == max([len(z) for z in x])][0]
                except:
                    return [(0,0,)]

            def get_definition_dict (page_tuple,page_dict):
                return_dict = {}
                for p in page_dict:
                    return_dict[p] = page_tuple[1]+(p-page_tuple[0])
                return return_dict

            def get_page_strings(page_dict):
                return_dict = {}
                lower_bound = min([page_dict[x] for x in page_dict])
                
                for p in list(page_dict.keys()):
                    if page_dict[p]<1:
                        return_dict[p] = int_to_roman(page_dict[p]-lower_bound+1)
                    else:
                        return_dict[p] = str(page_dict[p])
                return return_dict           
                     
            possible_pages = get_possible_pages(pdf)
            pages_one = get_page_lists(possible_pages,index=0)
            pages_two = get_page_lists(possible_pages,index=1)

            
            
            longest = find_longest_string(pages_one+pages_two)[-1]
            pages = get_definition_dict(longest,possible_pages)

            pages = get_page_strings(pages)

            return pages
            
            
            
        
             

                
                

            

        def add_capitalized (capitalized_words, last_cap_phrase='',page_no=0,position_at=0):

            # To add a capitalized phrase to the dictionary
            cap_phrase = ''

            if len (capitalized_words)>1:
                if capitalized_words[-1] in NAME_WORDS:
                    capitalized_words = capitalized_words[0:-1]
            while capitalized_words:
                
                if ((capitalized_words and
                     capitalized_words[0]
                     in self.first_words)):
                    capitalized_words = capitalized_words[1:]
                else:
                    break
                    
                
            if len(capitalized_words)>1:
                
                # join together capitalized words and add them to dictionary
                
                cap_phrase = ' '.join(capitalized_words).strip()
                cap_phrase = cap_phrase.replace('COPYRIGHT-PROTECTED MATERIAL','')
                cap_phrase = cap_phrase.replace('UNCORRECTED INDEXABLE PROOFS','')
                
                
            if "'s" in cap_phrase:
                cap_phrase, other_phrase = cap_phrase.split("'s")[0],''.join(cap_phrase.split("'s")[1:]).split(' ')
                
                add_capitalized(other_phrase,last_cap_phrase=last_cap_phrase,page_no=page_no,position_at=position_at)
                
            elif capitalized_words: 
                if exclude_first and not speller.unknown(set(capitalized_words[0].lower())):
                    self.first_words.add(capitalized_words[0])
                else:
                    cap_phrase = capitalized_words[0]
                    
            if cap_phrase:
                self.capitalized_phrases.add(cap_phrase)
                if cap_phrase not in self.order_dictionary:
                        self.order_dictionary[cap_phrase] = (page_no, position_at)
                        
                if cap_phrase in self.capital_histio:
                    self.capital_histio[cap_phrase].add(page_no)
                else:
                    self.capital_histio[cap_phrase] = {page_no}
                if ' ' in cap_phrase and len(cap_phrase)>15:
                    return cap_phrase
                return last_cap_phrase
            return last_cap_phrase

        

        with pdfplumber.open(directoryname+folder+filename) as pdf:
            initiated = False
            average_spacing = 0
            identified_pages = None
            char_list = []
            
            self.all_text = ''


            if self.database_object or yes_no_input('GET PAGES AUTOMATICALLY?'):
                
                if not self.above and not self.below:
                    above = yes_no_input('PAGE NUMBER ABOVE?')
                else:
                    above = self.above
                if not self.below:
                    below = not above or yes_no_input('PAGE NUMBER BELOW?')
                else:
                    below = self.below 
                identified_pages = determine_page_numeration (pdf,above=above,below=below)


            

            #Main procedure for analyzing the pdf and extracting text

            exclude_first = self.database_object or yes_no_input('Exclude single capital words if in spelling dictionary?')

            if not identified_pages and not self.database_object:
            
                from_this, to_that = proper_page(input('START FROM? ')), proper_page  (input('GO TO? '),to=True)
            else:
                from_this, to_that = 0,1000000
            
            from_this, to_that = get_limits(pdf,from_this,to_that)
             
            if not self.database_object and yes_no_input('REVIEW ALL FONTS IN PDF? '):
                excluded_fonts = determine_exclude_set(get_all_fonts_from_text(pdf,from_this,to_that))
            for pdf_page, page in enumerate(pdf.pages[from_this:to_that]):
                page_start = position_at
                
                
                def up_list(x,y):
                    for z in y:
                        x.append(z)
                
                text = page.extract_text()
                

                if not self.database_object:

                    try:

                        tokenized_text = word_tokenize(text)
                        words_with_tokens = nltk.pos_tag(tokenized_text)
                        all_nouns = set()

                        for w in words_with_tokens:

                            if w[1] in ['NN','NNS']:
                                all_nouns.add(w[0])
                        
                        self.nouns.update(all_nouns)
                    except:
                        print("TOKENIZER ERROR")
                    
                            
                
                
                query = False
                once_through = False
                exclude = False
                if not identified_pages:

                    while True:
                        page_no = extract_page_number(text,from_top=from_top).strip()
                        print(page_no)
                        if (not page_no.replace(' ','')
                            or page_no.isnumeric()
                            or page_no in first_romans):
                            break
                        if query:
                            print(text+'\n'+DIVIDER)
                            page_no = input('Page no ?')
                            query = False
                            if page_no.isnumeric():
                                break
                        
                            
                        if not once_through:
                            query = True
                            from_top = not from_top
                            once_through = True
                    if (page_no.isnumeric() and last_found_page
                        and last_found_page.isnumeric()
                        and int(page_no) == int(last_found_page)+1):
                        last_found_page = page_no
                    elif (initiated and last_found_page
                          and last_found_page.isnumeric()):
                        page_no = str(int(last_found_page)+1)
                        print('AUTOMATIC '+page_no)
                        last_found_page = page_no
                    else:
                        
                        print(text)
                        print(DIVIDER)
                        inp = input('Keep '+page_no+' or X to exclude PAGE, OR Q(uit) to STOP READING')
                        if not inp in YESTERMS + ['X','Q']:
                            page_no = input('Enter new page number?').strip()
                            if not initiated and page_no.isnumeric():
                                initiated = yes_no_input('Initiate automatic pagination?') 
                        if inp == 'X':
                            exclude = True
                        elif inp == 'Q':
                            break
                            
                        last_found_page = page_no
                            
                        if first_page:
                            inp = type_input(prompt='Keep '+page_no+' or X to exclude PAGE, OR Q(uit) to STOP READING ',must_be=['X','Q','Y',' '])
                            if not inp in YESTERMS + ['X','Q']:
                                page_no = input('Enter new page number?').strip()
                                if not initiated and page_no.isnumeric():
                                    initiated = yes_no_input('Initiate automatic pagination?')
                            if inp == 'X':
                                exclude = True
                            elif inp == 'Q':
                                 break
                            first_page = False
                            last_found_page = page_no
                
                if identified_pages or not exclude:

                    if identified_pages:
                        page_no = identified_pages[pdf_page]
                        print(pdf_page,page_no)

                    if self.in_italics:
                        
                        italic_pages.append(page_no)
                    
                    self.pages[page_no]=text
                    italics = ''

                    
                    
                    italic_found = False
                    last_size = 0
                    last_y0 = 0
                    hard_return = True
                    for char in page.chars:
                        
                        if char['y0']-last_y0>50:
                            last_y0=char['y0']
                            last_alpha_y0 = char['y0']
                            
                            

                        if char['y0']<300 and char['size'] != last_size and (abs(char['y0']-last_y0)>10)  and last_size!=0:
                            up_list(char_list,BREAK_PHRASE)
                            
##                            self.all_text+= BREAK_PHRASE
##                            position_at += BREAK_PHRASE_LEN
                        if (last_y0-char['y0'])>10: #CHANGES

##
##                            
##                            self.all_text+='\n'*abs(int((last_y0-char['y0'])/10))
                            up_list(char_list,'\n'*abs(int((last_y0-char['y0'])/10)))
                            
##                            position_at += abs(int((last_y0-char['y0'])/10))
##                            self.all_text+=' '*int(char['x0']/10)
                            up_list(char_list,' '*int(char['x0']/10))
##                            position_at += int(char['x0']/10)
                            hard_return = True 
                            
                        
                        if  char['size'] == last_size and char['y0']==last_y0 and (char['text']+last_char).isalpha():
                            new_spacing = char['x0']-last_x0
                            if new_spacing > average_spacing:
                                average_spacing = new_spacing
                        if (last_alpha_y0 == char['y0'] and (char['x0']-last_alpha_x0)>average_spacing):
                            if average_spacing > 0:
                                add_to = int((char['x0']-last_alpha_x0)/average_spacing)*' '
                                up_list(char_list,add_to)
                            
##                                
##                                self.all_text += add_to
##                                position_at += len(add_to)
##                           
                        
##                        print('LT',from_last_title,last_title)
                        if last_title:
                            from_last_title += 1
                        if last_cap_phrase:
                            from_last_cap_phrase += 1

                        if from_last_title > 200:
                            last_title = None
                            from_last_title = 0
                        if from_last_cap_phrase > 200:
                            last_cap_phrase = None
                            from_last_cap_phrase = 0
                        if char['text'] in ['’','’']: #convert single smartquote
                            char['text'] = "'"

                        if char['fontname'] not in excluded_fonts:
##                            self.all_text += char['text']
                            up_list(char_list,char['text'])
                            char_set.add(char['text'])

                            if self.in_italics:                   
                                if 'Italic' in char['fontname']:
                                    self.italicized_chars.add(page_start+len(char_list))
                                    
                                    if not italic_found and italics and italics[0].isupper():
                                        italics = italics.strip()
                                        if italics.endswith(')') and '(' in italics:
                                            italics = italics.split('(')[0].strip()
                                        italics = strip_punctuation(italics)
                                        
                                        self.italicized_phrases.add(italics)
                                        if italics not in self.order_dictionary:
                                            self.order_dictionary[italics] = (page_no,page_start+len(char_list))
                                        if italics not in self.italic_dict:    
                                            self.italic_dict[italics] = set()
                                        self.italic_dict[italics].add(page_no)
                                        if italics not in self.title_dict:
                                            self.title_dict[italics] = set()
                                        self.title_dict[italics].add(page_no)
                                        if italics not in self.title_translation_dictionary:
                                            self.title_translation_dictionary[italics]=''
                                            last_title = italics
                                        if italics not in self.author_dict:
                                            if last_cap_phrase:
                                                self.author_dict[italics] = last_cap_phrase
                                            
                                        italics = char['text']
                                        italic_found = True 
                                    else:
                                        italics += char['text']
                                        italic_found = True
                                else:
                                    if italic_found and char['text'] not in string.whitespace:
                                        italic_found = False

                            if self.in_parens:

                                if char['text'] == '(':
                                    found_left_paren = True
                                    
                                if char['text'] == '[':
                                    found_left_bracket = True

                            if self.in_quotes:

                                if char['text'] == '“':
                                    if not found_left_quote:
                                        found_left_quote = True       

    ##                                    
    ##                                else:
    ##                                    char['text'] == '”'
    ####                                    self.quoted_phrases.add(quoted_string)
    ####                                    quoted_string = ''
                                    char['text'] = '"'
                            
                            
                            if (char['text'].isupper() and last_char.islower())\
                               or char['text'] in string.punctuation+string.whitespace\
                               or (char['text'] in ['’',"'"] and last_char in string.punctuation+string.whitespace):
                                # WHEN A WORD IS FINISHED
                                oldword = word
                                word = de_hyphen(word)
                                simple_word = word.replace("'s",'')
                                if oldword!=word:
                                    self.dehyphenated[word] = oldword
                                    if self.in_italics:
                                        italics=italics.replace(oldword,word)
                                    if self.in_quotes:
                                        quoted_string=quoted_string.replace(oldword,word)
                                    if self.all_sentences:
                                        sentence_string=list(''.join(sentence_string).replace(oldword,word))


                                if self.database_object:
                                    if simple_word and simple_word != word:
                                        self.database_object.cursor.execute("INSERT OR REPLACE INTO words (book,word,page) VALUES (?,?,?);",(text_title,simple_word,page_no))
##                                        self.database_object.add_word(book=text_title,
##                                                                      word=simple_word,
##                                                                      page=page_no)
                                    if word:
##                                        self.database_object.add_word(book=text_title,
##                                                                          word=word,
##                                                                          page=page_no)
                                        self.database_object.cursor.execute("INSERT OR REPLACE INTO words (book,word,page) VALUES (?,?,?);",(text_title,word,page_no))
                                    
                                else:
                                    
                                    if simple_word in self.words:
                                        
                                        self.words[simple_word].add(page_no)
                                        self.histio[simple_word]+=1
                                        
                                    else:
                                        self.words[simple_word] = {page_no}
                                        self.histio[simple_word] = 1

                                if self.all_sentences:

                                    if self.database_object:

                                        self.database_object.cursor.execute("INSERT OR REPLACE INTO words_sentences (book,word,sentence) VALUES (?,?,?);",(text_title,simple_word,sentence_count))

                                    else:
                                        
                                        if word in self.words_in_sentences:
                                            self.words_in_sentences[word].add(sentence_count)
                                            
                                        else:
                                            self.words_in_sentences[word] = {sentence_count}
                                    
                                    
                                if word:

                                    if self.in_caps:
                                
                                        if (last_capitalized and  
                                            (((not word in NAME_WORDS and word[0].islower())
                                                and (not italic_found or add_italic_caps)) or
                                            ((char['text']!=',' and char['text'] in string.punctuation)
                                             or (char['text']==' ' and last_char==',')))):
                                                # To exclude the first word of a series of capital words
                                                # if it is either the first word of a sentence or has an apostrophe
                                            last_cap_phrase = add_capitalized (capitalized_words,last_cap_phrase,page_no,page_start+len(char_list))
                                            capitalized_words = []
                                            last_capitalized = False
                                            
                                            
                                                        
                                        elif (not word.isnumeric()) and (not italic_found or add_italic_caps) and (word[0].isupper() and not last_capitalized):
                                            capitalized_words.append(word)
                                            
                                            last_capitalized = True
                                        elif (not word.isnumeric()) and ((not italic_found or add_italic_caps) and (last_capitalized and (word[0].isupper or word in NAME_WORDS))):
                                            capitalized_words.append(word)
                                            

                                       
                                    word = ''
                                
                            else:
                                word+=char['text']

                            if self.all_sentences:


                                if char['text'] in ' ' and last_not_alpha:

                                    up_list(sentence_string,char['text'])

                                    if not self.database_object:
                                        

                                        self.sentence_record.append((sentence_starts_at,
                                                                     page_start+len(char_list),
                                                                     sentence_count))
                                        
                                        self.sentences.add(''.join(sentence_string))
                                        self.sentence_dict[sentence_count] = (sentence_starts_at,
                                                                              page_start+len(char_list),','.join(sentence_pages))
                                    else:

                                        self.database_object.cursor.execute("INSERT OR REPLACE INTO sentences (book,sentence,from_here,to_there) VALUES (?,?,?,?);",
                                                                            (text_title,sentence_count,sentence_starts_at,page_start+len(sentence_string)))
                                        self.database_object.cursor.execute("INSERT OR REPLACE INTO pages_for_sentences (book,sentence,page) VALUES (?,?,?);",
                                                                            (text_title,sentence_count,page_no))
                                        
                                    sentence_count += 1
                                    sentence_pages = set()
                                    sentence_string = []
                                    sentence_starts_at = page_start+len(char_list)
                                else:
                                    up_list(sentence_string,char['text'])
                                    sentence_pages.add(page_no)

                            if char['text'] not in string.whitespace+string.punctuation+'0123456789':
                                last_not_alpha = False
                            else:
                                last_not_alpha = True

                            if self.in_parens:
                                if found_left_paren:
                                    paren_string += char['text']
                                if found_left_bracket:
                                    
                                    bracket_string += char['text']
                            if self.in_quotes:
                                if found_left_quote:
                                    quoted_string += char['text']
                                

                            if self.in_parens:
                                if found_left_paren and char['text'] == ')':
                                    paren_string = paren_string[1:-1]
                                    
                                    self.parenthetical_phrases.add(paren_string)
                                    if ',' in paren_string and is_date(paren_string.split(',')[-1]):
                                        paren_string = ','.join(paren_string.split(',')[0:-1])
                                        
                                            
                                        
                                    if paren_string not in self.title_dict:
                                        self.title_dict[paren_string] = {page_no}
                                    else:
                                        self.title_dict[paren_string].add(page_no)
                                
                                    if last_title and self.title_translation_dictionary[last_title] == '':
                                        if .5 < len(paren_string)/len(last_title) < 2 and last_title != paren_string:
                                            self.title_translation_dictionary[last_title] = paren_string
                                            print(last_title,paren_string)
                                    
                                    
                                   
                                    found_left_paren = False
                                    paren_string = ''
                                if found_left_bracket and char['text'] == ']':
                                    bracket_string = bracket_string[1:-1]
                                    self.brackets.add(bracket_string)
                                    found_left_bracket = False
                                    bracket_string = ''
                            if self.in_quotes:
                                if found_left_quote and char['text'] == '”':
                                    char['text'] = '"'
                                    quoted_string = quoted_string[1:-1]
                                    if quoted_string[-1] in [',','.',';',':']:
                                        quoted_string = quoted_string[:-1]
                                    if quoted_string[-1] in ["'","’"]:
                                        if quoted_string[-2] in [',','.',';',':']:
                                            quoted_string = quoted_string[:-2]+quoted_string[-1]
                                        
                                    self.quoted_phrases.add(quoted_string)
                                    if quoted_string not in self.title_translation_dictionary:
                                         self.title_translation_dictionary[quoted_string]=''
                                         last_title = quoted_string
                                    if quoted_string not in self.title_dict:
                                        self.title_dict[quoted_string] = set()
                                    if quoted_string not in self.author_dict:
                                        if last_cap_phrase:
                                            self.author_dict[quoted_string] = last_cap_phrase
                                    self.title_dict[quoted_string].add(page_no)
                                    found_left_quote = False
                                    quoted_string = ''
          

                            last_char = char['text']
                            last_size = char['size']
                            last_y0 = char['y0']
                            last_x0 = char['x0']
                            if not char['text'].isnumeric():
                                last_alpha_x0 = char['x0']
                                last_alpha_y0 = char['y0']
                    page_text = ''.join(char_list)
                    page_len = len(char_list)

                    position_at += page_len
                    
                    char_list = []
                    self.all_text += page_text
                                    
                                    

                    if self.database_object:
                            self.database_object.add_page(book=text_title,page=page_no,from_here=page_start,to_there=position_at)
                    else:
                                
                        self.page_record.append((page_start,position_at,page_no))
                        self.page_dict[page_no]=(page_start,position_at)
            for word in self.first_words:
                if word and word[0].isupper() and not word.lower()+' ' in self.all_text:
                    self.capitalized_phrases.add(word)
            if self.database_object:
                self.database_object.add_text(book_name=text_title,
                              text=self.all_text)


                     
    def switch_text (self,text_title):

        if self.database_object and text_title:
            if text_title != self.last_working_title:
                self.last_working_title = text_title
                self.all_text = self.database_object.get_text(self.last_working_title)
        
    
    def print_section (self,from_point=0,to_point=0,italics=True,line_len=50,text_title=''):

        

        """Prints text from from_point to to_point.
        Line_len to determine the length of the line to be shown
        """

        def max_line_length(x):

            return max(len(x) for x in x.split('\n'))
        def split_long_line(x):
            
            first_half = x[0:int(len(x)/2)]
            second_half = x[int(len(x)/2):]
            for c in [' ','.',',',';',':','-']:
 
                if c in second_half:
                    first_half += second_half.split(c)[0]
                    second_half = ''.join(second_half.split(c)[1:])
                    break
            return first_half, second_half
        
        if self.database_object and text_title:
            if text_title != self.last_working_title:
                self.last_working_title = text_title
                self.all_text = self.database_object.get_text(self.last_working_title)
                
                

        
        text = self.all_text[from_point:to_point]

              
        if max_line_length(text)>200:

            return_text = ''
            line=''
        
            for char in text:
    ##            if char == '\n':
    ##                print(line)
    ##                print()
    ##                line = ''

                if len(line) > line_len:
                    if char == ' ':
                        return_text += line + '\n'
                        line = ''
                line += char
            return_text += line
            return return_text
        text = text.split('\n')
        temp_list = []                
        for x in text:                                                                                                    
            if len(x)>100:
                tx = split_long_line(x)
                temp_list.append(tx[0])
                temp_list.append(tx[1])
            else:
                temp_list.append(x)
        text = '\n'.join(temp_list)
        
        
        return text
            
    def get_from (self,index,obj=None,text_title=''):

        """Used for printing a section of an object
        containing a collection of different ordered text
        segments, such as the dictionary of pages.
        """

        if not self.database_object:
        
            if obj==None:
                obj=self.page_dict

            if index in obj:
                from_point,to_point = obj[index][0],obj[index][1]
                return self.print_section(from_point,to_point)

        from_point,to_point = self.database_object.get_page_range(book_name=text_title,
                                                                  page=index)
        return self.print_section(from_point,to_point,text_title=text_title)
        

    def get_page (self,page,text_title=''):

        """Returns the text of an entire page"""
        
        return self.get_from(page,obj=self.page_dict,text_title=text_title)
    def get_sentence (self,sentence):

        """The a sentence"""
        
        return self.get_from(sentence,obj=self.sentence_dict)
        
        
class Headings:

    """For establishing the titles of the index"""

    def __init__ (self,book_object=None):

        self.names = {'keep':set(),'purge':set(),'unprocessed':None}
        self.titles = {'keep':set(),'purge':set(),'ip':None,'qp':None,'pp1':None,'pp2':None}
        self.concepts = {'keep':set(),'purge':set()}
        self.book_object = book_object
        self.names_done = False
        self.titles_done = False
        self.italicized_phrases_done = False
        self.quoted_phrases_done = False
        self.paren_phrases1_done = False
        self.paren_phrases2_done = False 

    def get_names (self,name):

        """Takes a long name and divides it into its components
        """
        
        if ',' in name:
            last, first = name.split(',')[0].strip(),name.split(',')[1].strip()
        else:
            last, first = name.strip(),''
            
        return_names = set()
        for x in sorted(self.names['keep']):
            if ',' in x:
                if last in x.split(',')[0] and first in x.split(',')[1]:
                    return_names.add(x)
            else:
                if last in x:
                    return_names.add(x)
        return return_names


    def divide_set (self,entry_set,before,after,add_to=False,queries='krmnaxy',obj_type='',names=False):

        """
        Processes a set into various groups, and modifies forms of entries
        The basic format of an entry is:

        <AUTHOR>HEADING_SUBHEADING;SEARCH
        """

        return_keep = set()
        return_purge = set()
        input_phrase = ('RETURN to KEEP\n'*('k' in queries)+
                        'SPACE+RETURN to PURGE\n'*('r' in queries)+
                        '/ to MODIFY\n'*('m' in queries)+
                        '; to ADD SEARCH STRING\n'*('s' in queries)+
                        ': to CONVERT A NAME TO REGULAR FORM\n'*('n' in queries)+
                        '> to ADD AUTHOR\n'*('a' in queries)+
                        '_ to ADD SUBHEADINGS\n'*('x' in queries)+
                        '< to ADD AN ADDITIONAL HEADING\n'*('y' in queries)+
                        '^ to ADD A REFERENCE (see also)'*('z' in queries)+
                        '{ to ADD PAGES'+
                        'P to PURGE ALL'+
                        'Q to QUIT\n')
        
        go_on = True
        returned_entry_set = copy.copy(entry_set)
        new_x = ''
        inp = ''

        def get_related_for_phrase (term,entry_set):

            """Gets all the terms from entry_set that
            contain term"""

            def get_all_related (word):
                
                if not (word in SMALL_WORDS+NAME_WORDS+[' ']) and len(word)>5: 
                    return {x for x in entry_set 
                            if word in x}
                return set()

            return_set = set()
            for word in term.split(' '):
                return_set.update(get_all_related(word))
            if len(return_set)>10:
                return_set = set(sorted(return_set)[0:10])
            return return_set
        
        def is_contained (term, entryset):

            """True is some value of entryset contains, but is not identical, to term"""

            for x in entryset:
                if term != x:
                    if term.strip()+' ' in x+' ':
                        return True
            return False

            

        def get_order (x,page=False):

            """Returns a value indicating order of element in text"""

            if x in self.book_object.order_dictionary:
                if not page:
                    return self.book_object.order_dictionary[x][1]
                else:
                    return self.book_object.order_dictionary[x][0]
            elif x in self.book_object.all_text:
                return self.book_object.all_text.index(x)
            else:
                return 0
            

        def by_last (x):


            return x.split(' ')[-1]
        def dummy (x):
            return x
        while True:
            temp_inp = input('(1) SORT BY ORDER OF APPEARANCE?\n'+
                             ' 2) Sort alphabetically\n'+
                             '(3) Sort alphabetically by last word\n\n')
            if temp_inp in ['1','2','3']:
                break
        
        temp_fun = {'1':get_order,
                    '2':dummy,
                    '3':by_last}[temp_inp]

           
        iterate_over = sorted(entry_set,key=lambda z: temp_fun(z))
        new_list = []
        skipped_list = []
        for x in iterate_over:
            if is_contained(x,iterate_over):
                skipped_list.append(x)
            else:
                new_list.append(x)
        iterate_over = new_list
        print('SKIPPED')
        print(', '.join(skipped_list))
        
        total_length = str(len(iterate_over))


        page_at = None
        if yes_no_input('SHOW ALL?'):
            for counter, x in enumerate(iterate_over):
                print(counter,":",x,get_order(x))
        purge_all = False
        

        for counter, x in enumerate(iterate_over):
            
            x = x.strip()
            if x in self.book_object.order_dictionary:
                page_at = self.book_object.order_dictionary[x][0]

            if not inp == 'Q':
                returned_entry_set.discard(x)
 

            if add_to and not x.startswith(before):
                x = before+x
            if add_to and not x.endswith(after):
                x = x+after
            
            while go_on:


                x_reduced = x.replace(LEFT_ITALIC,'')\
                            .replace(RIGHT_ITALIC,'')\
                            .replace('“','')\
                            .replace('”','')\
                            .replace('"','')

                print('____________________________________________________________')
                if page_at:
                    print (str(counter)+'/'+total_length+' PAGE= '+page_at)
                    print('____________________________________________________________')
                
                
                print(before+after+'  '+x)
                print(input_phrase)
                print('**********RELATED PHRASES**********')
                for counter,xx in enumerate(get_related_for_phrase(x,iterate_over)):
                    print(str(counter+1)+' '*(10-len(str(counter+1)))+':'+xx)
                if x_reduced in self.book_object.title_translation_dictionary:
                    print('***********POSSIBLE TRANSLATION************')
                    print(self.book_object.title_translation_dictionary[x_reduced])
                if x_reduced in self.book_object.author_dict:
                    print('***********POSSIBLE AUTHOR*****************')
                    print(self.book_object.author_dict[x_reduced])
                if not purge_all:
                    inp = input('? ')


                
                if not inp and 'k' in queries:
                    return_keep.add(x)
                    print(x+' KEPT')
                    break

                elif inp == 'Q':
                    go_on = False
                    return_keep.add(x)

                elif not purge_all and inp == 'P':
                    if yes_no_input('ARE YOU SURE?'):
                        purge_all = True
                        
                
                elif inp == 'P' or purge_all or (inp == ' ' and 'r' in queries):
                    return_purge.add(x)
                    return_keep.discard(x)
                    print(x+' PURGED')
                    
                    break
                
                elif inp == '/' and 'm' in queries:
                    new = input(x)
                    if new:
                        new_x = new
                        
                elif inp == ';' and 's' in queries:
                    sh = input('ENTER SEARCHSTRING!')
                    if not input('\nRETURN TO ACCEPT\n'):
                        new_x = x+';;'+sh

                elif inp == '^' and 'z' in queries:
                    cross_reference = input('ENTER CROSSREFERENCE? ')
                    if yes_no_input('ACCEPT '+cross_reference):
                        if ';;' in x:
                            new_x = x.split(';;')[0]+'['+cross_reference+']'+';;'+x.split(';;')[1]
                        
                elif inp == ':' and 'n' in queries:
                    final_name = ''
                    all_names = x.replace("'s",'').replace('’s','').split(' ')
                    standard_name = all_names[-1]+', '+' '.join(all_names[0:-1])
                    for counter, name in enumerate(all_names):
                        print(counter,name)
                    
                    order = input('\n\tINPUT NAME for example: 01,23; 1von 0\n\tor RETURN FOR '+standard_name)
                    if not order:
                        final_name = standard_name
                    else:
                        for char in order:
                            print(char)
                            if char in '01234567809' and int(char) < len(all_names):
                                final_name += all_names[int(char)]
                                final_name += ' '
            
                            elif char == ',':
                                final_name += ', '
                            elif char:
                                final_name += char
                        final_name = final_name.strip().replace('  ',' ').replace(' ,',',')
                        
                    temp_inp = input('\n\tRETURN TO ACCEPT: '+final_name+'\nRETURN*2 TO ACCEPT + ADD NEW\n')
                    if not temp_inp or temp_inp=='  ':
                        new_x = final_name
                    if not temp_inp:
                        break
                            
                elif inp == '<' and 'y' in queries:
                    while True:
                        new_entry = input('\nENTER NEW HEAD ENTRY!')
                        if  new_entry and not input('RETURN TO ACCEPT'):
                            break
                    if new_entry:
                        return_keep.add(new_entry)
                elif inp == '>' and 'a' in queries:
                    if '<' in x and '>' in x:
                        x = x.split('<')[0] + x.split('>')[1]
                    name = input('Name of author for '+x+'\n')
                    possible_names = sorted(self.get_names (name))
                    for counter, pos_name in enumerate(possible_names):
                        print(counter,pos_name)
                    fulfilled = False
                    while True:
                        entered = input('ENTER NUMBER OR ? to ADD A NEW NAME')
                        if entered == '?':
                            while True:
                                new_name = input('ENTER NEW NAME!')
                                if new_name and not input('RETURN to ACCEPT!'):
                                    break
                            if new_name:
                                self.names['keep'].add(new_name)
                                new_x = '<'+new_name+'>'+x
                                break
                                
                               
                            
                        elif (entered.isnumeric() and
                              (int(entered) < len(possible_names))):
                            
                            if not input('RETURN TO KEEP: '+possible_names[int(entered)]):
                                new_x = '<'+possible_names[int(entered)]+'>'+x
                                break
                elif inp == '{':

                    pages, x = get_if (x,left='{',right='}')
                    if pages:
                        print('REPLACING '+pages)
                    temp_inp = ''
                    while not temp_inp:
                        temp_inp = input('PAGE range?')
                        try:
                            dummy = de_range(temp_inp)
                        except:
                            print('INVALID INPUT!')
                            temp_inp = ''
                    if temp_inp:
                        while True:
                            op_inp = input('+,-,= or Q')
                            if op_inp in ['+','-','=','Q']:
                                break
                    if not op_inp == 'Q':
                        x += '{'+op_inp+temp_inp+'}'
                    
                                            
                elif inp == '_' and 'x' in queries:
                    while input('SPACE + RETURN to ENTER A SUBHEADING!'):
                        subheading = input('SUBHEADING?')
                        if not input('RETURN TO ACCEPT '+subheading):
                            if '[' in x:
                                head,body = x.split('[')
                                new_x = head+'_'+subheading+'['+body
                            elif ';;' in x:
                                head, body = x.split(';;')
                                new_x = head+'_'+subheading+';;'+body
                            else:
                                new_x = x+'_'+subheading

                    
                if new_x and new_x != x and inp != '_':            
                    return_keep.add(new_x)
                    return_purge.add(x)
                    print(new_x+' ADDED')
                else:
                    if new_x and new_x != x:
                        return_keep.add(new_x)
                        print(new_x+' ADDED')
                    if inp != ':':
                        return_keep.add(x)
                        print(x+' KEPT')
                    else:
                        return_purge.add(x)
                        print(x +' PURGED')
                    
        if entry_set and not returned_entry_set and yes_no_input('FINISH FIRST STAGE? '):
            if obj_type == 'n':
                self.names_done = True
            if obj_type == 'ip':
                self.italicized_phrases_done = True
            if obj_type == 'qp':
                self.quoted_phrases_done = True
            if obj_type == 'pp1':
                self.paren_phrases1_done = True
            if obj_type == 'pp2':
                self.paren_phrases2_done= True
                    
                
                
                    

        return return_keep, return_purge, returned_entry_set
        
    def add (self,entered_set,obj=None,before='',after='',add_to=False,queries='krmna',obj_type='',names=False):

        """Divides the entered set and add the two parts into the set of terms to be kept,
        and the set of terms to be removed"""
        
        keep, discard,returned_entry_set = self.divide_set(entered_set,before,after,add_to,queries,obj_type=obj_type,names=names)
        obj['keep'].update(keep)
        obj['purge'].update(discard)
        if not obj_type or obj_type == 'n':
            obj['unprocessed'] = returned_entry_set
        else:
            obj[obj_type] = returned_entry_set
        return returned_entry_set

    def reform (self,obj=None,before='',after='',from_keep_to_purge=True,queries='kr'):

        """To revise a division that has already been made
        """

        if from_keep_to_purge:
            entered_set = obj['keep']
        else:
            entered_set = obj['purge']
        keep, discard, dummy = self.divide_set(entry_set=entered_set,
                                        before=before,
                                        after=after,
                                        add_to=False,
                                        queries=queries)
        print('keep',';'.join(keep))
        print('discard',';'.join(discard))
        if from_keep_to_purge:
            for x in discard:
                if x in obj['keep']:
                    obj['keep'].discard(x)
                    obj['purge'].add(x)
            for x in keep:
                print('keeping ',x)
                obj['keep'].add(x)
                
        else:
            for x in discard:
                if x in obj['purge']:
                    obj['purge'].discard(x)
                    obj['keep'].add(x)
                else:
                    obj['keep'].add(x)
              
                
    def run_reform (self,obj,
                    first_query='krmnxsyz',
                    second_query='kr',title=False,name=False):

        """Queries which of the "reformations" are to be made
        (from purge to keep, or keep to purge)
        and calls the reform function accordingly"""

        print('TO KEEP',len(obj['keep']))
        print('TO PURGE', len(obj['purge']))

        if yes_no_input('Prepurge to keep?'):
            scroll = Select(sort_function=lambda x:x.lower())
            
            
            obj['keep'] = set(scroll.scroll_through(obj['keep']))
        

        if obj['keep'] or obj['purge']:
            if yes_no_input('REVISE? '):
                print(DIVIDER)
                print(DIVIDER+'REVISING FROM KEEP TO PURGE'+'\n'+DIVIDER)
                self.reform(obj,
                            before='',
                            after='',
                            from_keep_to_purge=True,
                            queries=first_query)
                print()
                print(DIVIDER+'REVISING FROM PURGE TO KEEP'+'\n'+DIVIDER)
                self.reform(obj,
                            before='',
                            after='',
                            from_keep_to_purge=False,
                            queries=second_query)
            return True
        return False 

    def define_names (self):

        """Initiates the determination of the proper names
        that will be used in the index,
        or resumes the determination"""

        def split_set (x_set):

            """Returns a new set with all elements in the set not
            contains 's, or to the right or left of 's"""

            return_set = set()
            for x in x_set:

                if "'s" in x:
                    left, right = x.split("'s")[0],''.join(x.split("'s")[1:])
                    if left:
                        return_set.add(left)
                    if right:
                        return_set.add(right)
                else:
                    return_set.add(x)
            return return_set 
                        


        if not self.names_done:

            if not self.names['unprocessed']:

                

                self.add(split_set(self.book_object.capitalized_phrases),
                         obj=self.names,
                         before='',
                         after='',
                         queries='krmnxsyz',
                         obj_type='n',
                         names=True)
            else:
                self.add(self.names['unprocessed'],
                         obj=self.names,
                         before='',
                         after='',
                         queries='krmnxsyz',
                         obj_type='n',
                         names=True)
                
            
        else:
            self.run_reform(self.names)

    def define_titles (self):

        """Initiates the determination of the titles
        that will be used in the index,
        or resumes the determination"""

        print('TITLES')       

        if not self.titles_done:
            
            print('________TITLES IN ITALICS_________')

            if not self.italicized_phrases_done:

                if self.titles['ip'] is None:
                    op_set = self.book_object.italicized_phrases
                else:
                    op_set = self.titles['ip']
                if not op_set:
                    self.italicized_phrases_done = True

                self.add({x for x in op_set
                        if (x[0].isupper() and some_letters(x))},
                       obj=self.titles,
                       before=LEFT_ITALIC,
                       after=RIGHT_ITALIC,
                       add_to=True,
                       queries='krmaxsyz',
                       obj_type='ip')

                
            print('________TITLES IN QUOTATION MARKS_______')

            if not self.quoted_phrases_done:

                if self.titles['qp'] is None:
                    op_set = self.book_object.quoted_phrases
                else:
                    op_set = self.titles['qp']
                if not op_set:
                    self.quoted_phrases_done = True
                self.add({x for x in op_set
                          if (x[0].isupper() and len(x.split(' '))<max_title_length)},
                         obj=self.titles,
                         before='“',
                         after='”',
                         add_to=True,
                         queries='krmaxsyz',
                         obj_type='qp')
            print('_________TITLES IN PARENTHESES________')
            if not self.paren_phrases1_done:
                if self.titles['pp1'] is None:
                    op_set = self.book_object.parenthetical_phrases
                else:
                    op_set = self.titles['pp1']
                if not op_set:
                    self.paren_phrases1_done  = True 
                
                self.add({un_date(x) for x in op_set if (x[0].isupper() and ',' in x and not '“' in x and not '"' in x and last_is_date(x))},
                         obj=self.titles,
                         before=LEFT_ITALIC,
                         after=RIGHT_ITALIC,
                         add_to=True,
                         queries='krmaxsyz',
                         obj_type='pp1')

            if not self.paren_phrases2_done:

                if self.titles['pp2'] is None:
                    op_set = self.book_object.parenthetical_phrases
                else:
                    op_set = self.titles['pp2']
                if not op_set:
                    self.paren_phrases2_done  = True 
                    
                
                self.add({un_date(x) for x in op_set if (x[0].isupper() and ',' in x
                                                         and ('“' in x or
                                                              '"' in x)
                                                         and last_is_date(x))},
                         obj=self.titles,
                         before='',
                         after='',
                         add_to=True,
                         queries='krmaxsyz',
                         obj_type='pp2')
            if (self.italicized_phrases_done and
                self.quoted_phrases_done and
                self.paren_phrases1_done and
                self.paren_phrases2_done):
                 self.titles_done = True 
        else:
            print('REFORM')
            print('T',self.titles)

            self.run_reform(self.titles,
                            first_query='krmaxsyz',
                            second_query='kr')
            

    def define_concepts (self):

        """Initiates the determination of the concepts
        that will be used in the index,
        or resumes the determination"""

        print('CONCEPTS')

        if not self.run_reform(self.concepts,
                               first_query='krmnxsz',
                               second_query='kr'):

            while True:
                temp_inp = input('(A)dd terms individually or (S)elect from text.')
                if temp_inp  in ['A','S']:
                    break
            if temp_inp == 'S':
                
                terms = self.book_object.nouns
                print(len(terms))
                terms = [x for x in terms if not x.isnumeric()]
                
                if yes_no_input('Purge frequent?'):
                    
                    terms = [x for x in terms if x not in English_frequent_words]
                print(len(terms))
                if yes_no_input('Purge mispelled?'):
                    
                    unknown = speller.unknown(terms)
                    terms = [x for x in terms if x not in unknown]
                print(len(terms))
                if yes_no_input('Pure capital words?'):
                    terms = [x for x in terms if not x[0].isupper()]

                
                while True:

                    while True:
                        lower, upper = None, None

                        lower = input('LOWER THRESHOLD?')
                        if lower.isnumeric():
                            lower = int(lower)
                        upper = input('UPPER THRESHOLD?')
                        if upper.isnumeric():
                            upper = int(upper)
                        if lower and upper:
                            break
                    terms = [x for x in terms if x in self.book_object.words and lower<=len(self.book_object.words[x])<=upper]
                    print('There are '+str(len(terms))+' terms in your set')
                    if yes_no_input('Show?'):
                        print(', '.join(terms))
                    if yes_no_input('Accept?'):
                        break
                self.add(set(terms),
                         self.concepts,
                         queries='krmnx')
                if yes_no_input('Add additional terms?'):
                    new_terms = [x.strip() for x in input('?').split(',')]
                    print (', '.join(new_terms))
                    if yes_no_input('ADD to concepts?'):
                        self.add(set(new_terms),
                                 self.concepts,
                                 queries='krmnxsz')
            else:
                self.add(set(),self.concepts,queries='krmnxsz')
            
            
        
        
class Searcher:

    """Find headings in the interpreted PDF text
    """

    def __init__ (self,book_object=None):

        self.book_object = book_object



    def convert_sentences_to_pages (self,return_pages,sentence_search=None,database_object=None,text_title=None):


        def twine (x,y):

            def conv (x):
                if isinstance(x,set):
                    return sorted(list(x),key=lambda x: page_sort_function(x))
                return x
            x,y = conv(x), conv(y)

            returnlist = []
            for c in range(min([len(x),len(y)])):
                returnlist.append((x[c],y[c]))
            return returnlist

        def untwine (x):

            returnlist_one = []
            returnlist_two = []
            
            for t in x:
                returnlist_one.append(t[0])
                returnlist_two.append(t[1])
            return returnlist_one,returnlist_two
                
        def reduce_twine (x,y):

           returnlist = []
           for yy in y:

               if yy[0] in x:
                   returnlist.append(yy)
           return returnlist



        """Converting sentences to pages"""

        if sentence_search or (return_pages and isinstance(list(return_pages)[0],int)):

            if not database_object:
                return_sentences = return_pages
                return_pages[t_t] = self.book_object.get_pages_for_sentences([int(x) for x in return_pages],text_title=text_title)
                sentence_page_pairs = twine(return_pages,return_sentences)
                return_sentences = untwine(reduce_twine(return_pages,sentence_page_pairs))[1]
                return_pages = set(return_sentences)
    
            else:

                for t_t in return_pages:
                    return_sentences = return_pages[t_t]
                   
                    pages_found = set()
                    for s in return_sentences:
                        database_object.cursor.execute("SELECT * FROM pages_for_sentences WHERE book=? AND sentence=?",(t_t,s,))
                        x = database_object.cursor.fetchone()[-1]
                        pages_found.add(x)                   

                    
                    return_pages[t_t] = pages_found
                    
        return return_pages
    
    

    def get_pages (self,term,dictionary_object=None,alternative_object=None,literal_object=None,database_object=None,text_title=None,sentence_search=False,get_terms_only=False):

        """Core search routine. Retrieves pages for the given term.
        """


        if term.startswith('~'):
            negative = True
            term = term[1:]
        else:
            negative = False
            
        count_min = None
        count_max = None
        count_equals = []
        count_not_equals = []
        
        must_not_be = set()
        must_be = set()

        qualifiers, term = get_if (term,'[',']',get_all = True)

        for qualifier in qualifiers:
            if qualifier.startswith('#'):
                qualifier = qualifier[1:]
                if qualifier:
                    if qualifier[0] in ['=','~']:
                        working_list = {'=':count_equals,
                                        '~':count_not_equals}[qualifier[0]]
                        qualifier = qualifier[1:]
                        for q in de_range(qualifier):
                            working_list.append(int(q))
                    elif qualifier[0] == '>':
                        count_min = int(qualifier[1:])
                    elif qualifier[0] == '<':
                        count_max = int(qualifier[1:])
            elif qualifier.startswith('+'):
                must_be = de_range(qualifier[1:])
            elif qualifier.startswith('-'):
                must_not_be = de_range(qualifier[1:])

        
            
            

        if term.startswith('_') and not database_object:
            dictionary_object = alternative_object
            term = term[1:]
        result = set()

        full_words = []
        center_words = []
        left_of = []
        right_of = []
        literal_search = False

        def interpret_caps(enterlist):

            #TO interpret CAPPHRASES 

            returnlist=[]
            for x in enterlist:
                if not x.isupper():
                    returnlist.append(x)
                else:
                    returnlist.append(x.lower())
                    returnlist.append(x[0].upper()+x[1:].lower())
            return returnlist 
        if term.count('/') == 1:
            if term.split('/')[1]:
                right_of.append(term.split('/')[1])
            elif term.split('/')[0]:
                left_of.append(term.split('/')[0])
        elif term.count('/') == 2 and term.split('/')[1]:
            center_words.append(term.split('/')[1])
        elif not term.startswith('$') and '/' not in term:
            full_words.append(term)
        elif term.startswith('$'):
            term = term[1:]
            center_words = [x for x in term.split('$') if x]
            literal_search = True 
             
        full_words = interpret_caps(full_words)
        center_words = interpret_caps(center_words)
        left_of = interpret_caps(left_of)
        right_of = interpret_caps(right_of)


        temp_terms = []
        return_pages = set()
        all_terms = []
        if database_object:
            return_pages = {}
            all_terms = {}
            

        if not literal_search:



            all_words = center_words+left_of+right_of

            temp_terms = []
                            
            if all_words:

                for word in all_words:

                    if not database_object:

                        words_from_dictionary = dictionary_object.keys()
##                    else:
##
##                        words_from_dictionary = database_object.get_all_words()

                    if center_words:

                        if not database_object:
                        
                            temp_terms = [x for x in words_from_dictionary if word in x]
                        else:
                            temp_terms = database_object.get_words_left_center_right(center=word)
                    elif left_of:

                        if not database_object: 
                            temp_terms = [x for x in words_from_dictionary if x.startswith(word)]
                        else:
                            temp_terms = database_object.get_words_left_center_right(left=word)
                    elif right_of:
                        if not database_object:
                            temp_terms = [x for x in words_from_dictionary if x.endswith(word)]
                        else:
                            temp_terms = database_object.get_words_left_center_right(right=word)

            

            if get_terms_only:
                return full_words+temp_terms

            
                        
                    

            for word in full_words+temp_terms:

                    if not database_object:

                        if word in dictionary_object:
                            return_pages = return_pages.union(dictionary_object[word])
                            all_terms.append(word)
                    else:
                        if isinstance(text_title,str):
                            if not text_title in return_pages:
                                return_pages[text_title] = set()
                            if not text_title in all_terms:
                                all_terms[text_title] = []
                            if not sentence_search:
                                if database_object.word_exists (text_title,word):
                                    return_pages[text_title] = return_pages[text_title].union(database_object.get_pages_for_word(text_title,word))
                                    all_terms[text_title].append(word)
                            else:
                                if database_object.word_exists (text_title,word):
                                    return_pages[text_title] = return_pages[text_title].union(database_object.get_sentences_for_word(text_title,word))
                                    all_terms[text_title].append(word)
                                                                

                        else:
                            if isinstance(text_title,list):
                                search_range = text_title
                            else:
                                search_range = database_object.get_books()

                            for text_name in search_range:

                                if not text_name in return_pages:
                                    return_pages[text_name] = set()
                                if not text_name in all_terms:
                                    all_terms[text_name] = []
                                words = database_object.get_words(text_name)
                                if not words:
                                    words = []

                                if not sentence_search:
                                    if database_object.word_exists (text_name,word):
                                        return_pages[text_name] = return_pages[text_name].union(database_object.get_pages_for_word(text_name,word))
                                        all_terms[text_name].append(word)
                                else:
                                    if database_object.word_exists (text_name,word):
                                        return_pages[text_name] = return_pages[text_name].union(database_object.get_sentences_for_word(text_name,word))
                                        all_terms[text_name].append(word)
                                        
                            
                            
                
                
        elif not database_object:
            
            for page in literal_object.keys():
                is_found = True 
                term_counter = 0
                page_text = literal_object[page]
                while is_found and term_counter<len(center_words):
                    term = center_words[term_counter]
                    if term in page_text:
                        page_text = page_text.split(term)[1]
                        term_counter += 1
                    else:
                        is_found = False
                if is_found:
                    
                    if len(center_words) == 1 and (count_min or count_max or count_equals or count_not_equals):
                        is_really_found = True
                        number_of = literal_object[page].count(center_words[0])
                        if count_min:
                            if number_of < count_min:
                                is_really_found = False
                        if count_max:
                            if number_of > count_max:
                                is_really_found = False
                        if count_equals:
                            if number_of not in count_equals:
                                is_really_found = False
                        if count_not_equals:
                            if number_of in count_not_equals:
                                is_really_found = False
                        if is_really_found:
                            return_pages.add(page)
                            for t in center_words:
                                all_terms.append(t)
                    else:
                        return_pages.add(page)
                        for t in center_words:
                            all_terms.append(t)

                
        if not database_object:        
            if negative:
                return_pages = {x for x in self.book_object.pages.keys() if x not in return_pages}
            return_pages = {x for x in return_pages
                            if (not must_be or x in must_be)
                            and (not must_not_be or x not in must_not_be)}


        else:
            for text_name in return_pages:

                

                if negative:
                    return_pages[text_name] = {x for x in self.book_object.pages.keys()
                                               if x not in return_pages[text_name]}
                return_pages[text_name] = {x for x in return_pages[text_name]
                                           if (not must_be or x in must_be)
                                           and (not must_not_be or x not in must_not_be)}


        return return_pages, all_terms
        

    def search_entry(self,term,dictionary_object=None,
                     alternative_object=None,
                     literal_object=None,
                     text_title=None,
                     database_object=None,
                     sentence_search=False):

        """Interprets a search term and runs the search accordingly


        THE SEQUENCE IS
            SEARCH_ENTRY =>
            SEARCH =>
            GET_PAGE """
        
        term = term.replace('<<','##LT##').replace('>>','##GT##')
        
        term = term.replace('’',"'")        
        if '<' in term:
            term = term.split('<')[1].split('>')[0]
        term = term.replace('##LT##','<').replace('##GT##','>')
        
        
        if ';;' in term:
            head, term = term.split(';;')[0], term.split(';;')[1]
            term = term.replace('%',head)
        term = term.replace(LEFT_ITALIC,'').replace(RIGHT_ITALIC,'')
        term = term.replace('“','').replace('”','').replace('"','')
        x = self.search (term,dictionary_object=dictionary_object,alternative_object=alternative_object,literal_object=literal_object,
                         text_title=text_title,database_object=database_object,
                         sentence_search=sentence_search)

         
        return self.convert_sentences_to_pages(x[0],sentence_search=sentence_search,database_object=database_object,text_title=text_title),x[1]

    def search (self,search_phrase,dictionary_object,alternative_object=None,literal_object=None,sentence_object=None,text_title=None,database_object=None,sentence_search=False):

        """Searches a complex search phrase
        ---allowing &=AND, |=OR, nested parentheses ---
        in the disctionary object.
        Uses the parser."""

        
        

        def get_terms (entry):

            """Extracts individual terms from search phrase"""

            for char in '()&|':
                entry = entry.replace(char,'[#]')
            all_terms = [x.strip() for x in entry.split('[#]') if x]
            return all_terms

        universe = {}

        all_terms = get_terms (search_phrase)
        parsed_phrase = parser.parse(search_phrase)
        found_terms = set()

        if not database_object:
            for term in all_terms:
                x = self.get_pages(term,dictionary_object=dictionary_object,
                                   alternative_object=alternative_object,
                                   literal_object=literal_object,
                                   text_title=text_title,
                                   sentence_search=sentence_search)
                universe[term] = x[0]
                found_terms.update(set(x[1]))
                
            result = parser.interpret(parsed_phrase,universe)
            return result, found_terms

        if isinstance(text_title,str):
            all_titles = [text_title]
        elif isinstance(text_title,(list,set)):
            all_titles = list(text_title)
            
        else:
            all_titles = database_object.get_books()

        result = {}
        found_terms = {}

        found_books = set()

        complete_terms = []
        for temp_term in get_terms(search_phrase):
            
            complete_terms += self.get_pages(temp_term,dictionary_object=dictionary_object,
                                       alternative_object=alternative_object,literal_object=literal_object,
                                       text_title=text_title,database_object=database_object,
                                       sentence_search=sentence_search,get_terms_only=True)
        for term in complete_terms:
            
            books = database_object.get_books_for_word(term)
            found_books.update(books)
        
            
        
        for title in set(all_titles).intersection(found_books):

            universe[title] = {}

            for term in all_terms:
                x = self.get_pages(term,dictionary_object=dictionary_object,
                                   alternative_object=alternative_object,literal_object=literal_object,
                                   text_title=title,database_object=database_object,
                                   sentence_search=sentence_search)
                if title in x[0]:
                    universe[title][term] = x[0][title]
                else:
                    universe[title][term] = set()
                
                if not title in found_terms:
                    found_terms [title] = set()
                if title in x[1]:
                    
                    found_terms[title].update(set(x[1][title]))
                
            result[title] = parser.interpret(parsed_phrase,universe[title])
        
        return result, found_terms
            

        
            

class Index_Maker:

    """Basic class for making indexes, once the text has been interpreted, and
        the list of entries has been determined"""
    

    def __init__ (self):

        self.project_name = ''
        self.text_filename = ''
        self.project_file = None
        self.text_object = None 
        self.names = False
        self.titles = False
        self.concepts = False
        self.index = None
        self.index_name = ''
        self.index_made = False
        self.index_text = ''
        self.reverse_index_text = ''
        self.project_object = None
        self.searcher_object = None
        self.reverse_table = {}
        self.index = {'names':{},
                      'titles':{},
                      'concepts':{}}
        self.return_dict = {}
        self.override = True


           


    def reverse_index (self):

        """Reverses the index, such that the keys are pages,
        and the values are the indexed terms"""

        if self.index:

            for index_type in self.index:
                for entry in self.index[index_type]:
                    typed_entry = entry + ' ('+index_type+')'
                    for page in self.index[index_type][entry]:
                        if page not in self.reverse_table:
                            self.reverse_table[page] = {typed_entry}
                        else:
                            self.reverse_table[page].add(typed_entry)

    def input_term (self,inp=None):

        menu = """
(0) LOAD TEXT %%
(1) OPEN PROJECT %
(2) SAVE PROJECT AS 
(3) SAVE PROJECT
(4) INITIATE PROJECT
(5) DEFINE NAMES %%%
(6) DEFINE TITLES %%%%
(7) DEFINE CONCEPTS %%%%%
(8) FIND PAGES %%%%%%
(9) DISPLAY INDEX
(10) DISPLAY ALL TERMS
(11) DISPLAY ALL PURGED TERMS
(12) FORMAT INDEX 
(13) EXAMINE PAGES
(14) CREATE REVERSE INDEX
(15) SHOW REVERSE INDEX
(16) CORRECT INDEX
(17) SAVE INDEX
(18) LOAD INDEX
(19) SAVE INDEX AS TEXT FILE 
(20) SEARCH TEXT BY PAGES
(21) SEARCH TEXT BY SENTENCES/ SHOW PAGES
(22) SEARCH TEXT BY SENTENCES/ SHOW SENTENCES
(23) OVERRIDE ERROR EXCEPTION
(24) RETURN INDEX
(25) QUIT
(26) SAVE AS TEXT
(27) LOAD AS TEXT
(28) CHANGE STATUS
(29) SHOW DEHYPHENATED WORDS
(30) SHOW FIRST WORDS
(31) PRE-PURGE NAMES
(32) LOAD FROM EXCEL TO NAMES
(33) LOAD FROM EXCEL TO TITLES
(34) LOAD FROM EXCEL TO CONCEPTS
(35) MULTISEARCH


"""
 
        
        def bracket(text,terms):

            """Surrounds all terms in text with brackets..."""
            return_set = set()

            for word in terms:
                if word in text:
                    text = text.replace(word,'<<'+word+'>>')\
                           .replace('<<<<','<<').replace('>>>>','>>')

                    return_set.add(word)
                    
                
            text = text.replace('>> <<',' ').replace('>><<','')
            return text,return_set

        def side_marks (text):

            """Adds sidemarks indicating found elements"""

                
            
            returnlines = []            

            for section in text.split(BREAK_PHRASE):

                returnlines.append('')
                
                
                lines = section.split('\n')
                max_line_length = max([len(x) for x in lines])
                for line in lines:
                    fragments,dummy = get_if(line,'<<','>>',True)
                    if '<<' in line:
                        line = '  ===>>> ' + line + (max_line_length-len(line))*' '+'    <<'+','.join(fragments)+'>>'
                    else:
                        line = '         ' + line
                    returnlines.append(line)
            return '\n'.join(returnlines)

        def recommend_classifiers(text,classifier_dict):

            """Recommends appropriate index subterms for a given page of text"""
            
            surround = lambda a:' '+a+' '
            recommendations = set()
            for x in classifier_dict.keys():

                if '/' in x:

                    classifier, search_term = x.split('/')[0],x.split('/')[1]
                else:
                    classifier = purge_small(x, SUBHEAD_WORDS)
                    search_term = classifier
                search_terms = search_term.split(' ')
                
                
                for y in search_terms:
                    if surround(y) in surround(text):
                        recommendations.add(x)
            return recommendations
                    
           

        def abridge_page (text,around=3):


            """Keeps only lines of page
            surrounding found terms"""
            
            text = text.split('\n')
            side_marked = set()
            to_keep = set()
            return_list = []
            for line_number, line in enumerate(text):
                if '===>>>' in line:
                    side_marked.add(line_number)
            for line in side_marked:
                for x in range(line-around,line+around+2):
                    to_keep.add(x)
            for line in sorted(to_keep):
                if 0 <= line < len(text):
                    return_list.append(text[line])
            return '\n'.join(return_list)
            
                    

        def pre_purge (iterate_over=None):

            """For a preliminary sifting of a set of items"""
            
        
            four_or_more_words = [x for x in iterate_over if x.count(' ')>=3]
            three_words = [x for x in iterate_over if x.count(' ')==2]
            two_words = [x for x in iterate_over if x.count(' ')==1]
            one_word = [x for x in iterate_over if x.count(' ')==0]


            scroll = Select(sort_function=lambda x:x.lower())
            four_or_more_words = set(scroll.scroll_through(four_or_more_words))
            three_words = set(scroll.scroll_through(three_words))
            two_words = set(scroll.scroll_through(two_words))
            one_word = set(scroll.scroll_through(one_word))
            word_set = one_word.union(two_words).union(three_words).union(four_or_more_words)
            return word_set


        def read_excel (filename=''):

            """To read an excel file containing terms to index.
            TERM;;OPT SEARCH PHRASE{OPT PAGES} \t SUBTERM1;SUBTERM2;SUBTERM3
            """

            def check(x,check_for=['()','{}','[]','<>']):
                for cf in check_for:
                    if x.count(cf[0])!=x.count(cf[1]):
                        print(x,' IS IRREGULAR WITH RESPECT TO ',cf)
                return x



            try:

                wb = load_workbook(directoryname+index_folder+filename+'.xlsx')
                ws = wb[wb.sheetnames[0]]
            except:

                print('FILE ERROR: CANNOT LOAD '+filename)


            if ws:
                while True:
                    min_row = input('Start from ROW?')
                    max_row = input('Go to ROW?')
                    if (min_row.isnumeric() or not min_row) or (max_row.isnumeric() or not max_row):
                        break
                if min_row:
                    min_row = int(min_row)
                if max_row:
                    max_row = int(max_row)
                all_entries = []
                if min_row:
                    y=min_row
                else:
                    y=1
                
                while True:
                    
                    first_cell = ws.cell(row=y,column=1)
                    second_cell = ws.cell(row=y,column=2)
                    if first_cell:
                        first_cell = first_cell.value
                    if not first_cell or (max_row and y>max_row):
                        break
                     
                    if second_cell:
                        second_cell = second_cell.value
                
                    if first_cell:
                        all_entries.append(check(first_cell.strip()))
                        head_search = ''
                        if ';;' in first_cell:
                            first_cell, head_search = first_cell.split(';;')[0],first_cell.split(';;')[1]
                        else:
                            head_search = '&'.join([x.strip() for x in first_cell.split(',')[0].split(' ') if x])
                        
                    if second_cell:
                        
                        
                        for sub_phrase in second_cell.strip().split(';'):
                            sub_phrase = sub_phrase.strip()
                            if sub_phrase:
                                search_phrase,sub_phrase = get_if(sub_phrase,left='<',right='>')
                                heading = first_cell.strip()+'_'+sub_phrase
                                if search_phrase:
                                    if not head_search:
                                        heading+=';;'+search_phrase
                                    else:
                                        heading+=';;'+head_search+'&'+'('+search_phrase+')'
                                all_entries.append(check(heading))
                    y += 1
                return all_entries
            return []
                       

        def run_search (inp='20',search_term=None,quit_immediately=False,show_reversed=False,database_object=None,text_title=None):

            """Runs the search of time inp
            """

            while True:

                classifier_phrase = ''
                formatted = ''
        
                if inp == '20':
                    obj = self.text_object.words
                    alt_obj = self.text_object.title_dict
                    temp_ob = 'concepts'
                    lit_obj = self.text_object.pages
                    
                    
                else:
                    obj = self.text_object.words_in_sentences
                    alt_obj = None
                    lit_obj = None
                if not self.searcher_object:
                    self.searcher_object = Searcher(self.text_object)
                if not search_term:
                    search_term = input('SEARCH TERM or Q(uit) ')
                if not search_term.strip() or search_term == 'Q':
                    if search_term == 'Q':
                        inp = ''
                    break
##                
##                if inp == '20' and search_term.startswith('_'):
##                    search_term = search_term[1:]
##                    obj = self.text_object.title_dict


                if '\t' in search_term:
                    search_term, classifier_phrase = search_term.split('\t')[0:2]
                x = self.searcher_object.search_entry(search_term,obj, alt_obj, lit_obj,database_object=None,text_title=None)
                results,terms = x[0], x[1]

                if inp == '20':

                    print('RESULTS: ',format_range(results))
                else:
                    print('RESULTS: ',', '.join(sorted([str(x)+'/'+self.text_object.sentence_dict[x][2] for x in results])))           


                
                print('TERMS: ',', '.join(terms))
                if results:
                    page_set, page_dict, rej_pages =None, None, None 
                    tt_inp = input('SHOW or (A)ll or (S)elect? ')
                    if tt_inp in YESTERMS+['A','S']:
                        
                        if inp == '21':
                            page_set,page_dict,rej_pages = show_pages (sorted(list(self.text_object.get_pages_for_sentences(results)),
                                                                 key=page_sort_function),
                                                          terms=terms,show_all=(tt_inp=='A' or tt_inp=='S'),
                                                          select=(tt_inp=='S'),abridge=(tt_inp=='A' or tt_inp=='S'),classifier_phrase=classifier_phrase,
                                                                       show_reversed=show_reversed)
                        elif inp == '22':
                            
                            page_set,page_dict,rej_pages = show_pages (sorted(list(results),key=page_sort_function),
                                                          sentences=True,
                                                          terms=terms,
                                                          show_all=(tt_inp=='A'or tt_inp=='S'),
                                                          select=(tt_inp=='S'),classifier_phrase=classifier_phrase,
                                                                       show_reversed=show_reversed)
                        else:
                            page_set,page_dict,rej_pages = show_pages(sorted(list(results),
                                              key=page_sort_function),
                                                         terms=terms,
                                                         show_all=(tt_inp=='A' or tt_inp=='S'),
                                                         abridge=(tt_inp=='A' or tt_inp=='S'),
                                                         select=(tt_inp=='S'),classifier_phrase=classifier_phrase,
                                                                       show_reversed=show_reversed)
                    print('FINAL RESULTS FOR '+','.join(terms))
                    print()
                    print()
                    formatted = format_result(search_term=search_term,page_set=page_set,page_dict=page_dict,rejected=rej_pages)
                    print(formatted)
                    print()
                    
                search_term=None
                if quit_immediately:
                    break
            return inp, formatted                  


        def run_multi_search (x,all_results=None,inp=inp):

            if not all_results:
                all_results = []

            go_on = True
            while go_on:
                if not skip_predetermined or not get_if(x,'{','}')[0]:
                    
                    switched = False # FOr when a new search term is used
                    if '\t' in x:
                        head, body = x.split('\t')[0:2]
                    else:
                        head = x
                        body = ''
                    


                    
                    if ';;' in head:
                        head_term, search_term = head.split(';;')[0:2]
                    else:
                        head_term, search_term = head, head

                        if '<' in head and '>' in head:
                            author, head_term = get_if(head,'<','>')

                            head_term, search_term = '<'+author+'>'+head_term, '_'+head_term
                        else:
                            
                            if ';' in search_term:
                                search_term = search_term.split(';')[0]
                            if ',' in search_term:
                                search_term = search_term.split(',')[0]
                            parenthetical,search_term = get_if(search_term,left='(',right=')')
                            if parenthetical == 's':
                                search_terms = search_term+'|'+search_term+'s'
                    
                        
                    t_inp = input('RETURN TO USE '+search_term+' OR ENTER NEW')
                    if t_inp.strip():
                        search_term = t_inp
                        switched = True
                    try:
                        empty, temp_result = run_search(inp,search_term=search_term,quit_immediately=True)
                    except:
                        print('SEARCH FAILED')
                        temp_result = ''
                    temp_result = head_term + temp_result
                    
                        
                    print('RESULT =',temp_result)
                    while True:
                        temp_inp = type_input('[K]eep, Add [T]erm, [P}ass, [E]dit, [S]witch mode, [R}erun, [Q]uit_and_keep or (A)bort',must_be=['K',' ','T','R','Q','A','P','E','S'],return_empty=False)
                        if not temp_inp in ['E','S','T']:
                            break
                        else:
                            if temp_inp == 'E':
                                while True:
                                    print(temp_result)
                                    tt_result = input('?')
                                    ttt_inp = type_input('Accept or (B)reak',must_be=YESTERMS+['B'],return_empty=True)
                                    if ttt_inp:
                                        if not ttt_inp == 'B':
                                            temp_result = tt_result
                                        break
                            elif temp_inp == 'T':
                                while True:
                                    new_term = input('?')
                                    if new_term:
                                        break
                                if new_term.strip():
                                    ttt_inp = yes_no_input('Accept '+new_term+' or RETURN')
                                    if ttt_inp:
                                        term_stack.add(new_term)
                                   
                                        
                            else:
                                inp = int(numeric_input('Enter search mode.\n1) Pages\n(2) Pages/sentences\n(3) Sentences\n',lower=1,upper=3)) + 19
                                inp = str(inp)
                                print('NEW MODE =',inp)
                                
                                
                                        
                            
                    if temp_inp in ['K','Q','P']:
                        if not temp_inp == 'P':
                            all_results.append(temp_result)
                        go_on = False
                    elif temp_inp == 'A':
                        temp_inp = 'Q'
                        go_on = False
                    print('ALL RESULTS')
                    print(DIVIDER)
                    print('\n'.join(all_results))
                    print(DIVIDER)
                    if temp_inp == 'Q':
                          return False, all_results
                elif add_automatically or yes_no_input('ADD '+x):
                    all_results.append(x)
                    go_on = False
                else:
                    print(x,'SKIPPED')
                    go_on = False
            
            return True, all_results


        def reform_term (x,obj_name='names'):


##            def uncurl(x):
##                return unpara(x,'{','}')

            """Takes the entire index term,
            and returns the appropriate search phrases,
            which is either all the words of the term joined together with the "&"
            or is a separate search phrase. If the latter includes %%,
            then this is replaced with the AND-form of the regular term.

            e.g. Frog life;;(%%) | Toad

            => (Frog & Life) | Toad
            
            """
            def unpara(x,left='(',right=')'):

                """Removes parentheses"""
                
                if not '(' in x or not ')' in x:
                    return(x)
                else:
                    return x.split('(')[0]+x.split(')')[1]
                
            search_phrase = None
            all_terms = []
##            x = uncurl(x)
            if '>' in x:
                x = x.split('>')[1]
            if '[' in x:
                x = x.split('[')[0]
            if ';;' in x:
                heading, search_phrase = x.split(';;')[0], x.split(';;')[1]
            else:
                heading = unpara(x)

            if obj_name in ['names','concepts']:
                if ',' in heading:
                    heading, body = heading.split(',')[0],heading.split(',')[1:]
                    all_terms = [heading] + body
            if not all_terms:
                if '_' in heading:
                    if heading.split('_')[1].strip().startswith('-'):
                        heading = heading.split('_')[1].strip()+heading.split('_')[0].strip()
                    else:
                        head_phrase = heading.split('_')[0]
                        sub_phrase = heading.split('_')[1]
                        sub_phrase = purge_small(sub_phrase, SUBHEAD_WORDS)
                        sub_phrase_terms = [x for x in sub_phrase.split(' ') if x]
                        heading = '&'.join([head_phrase]+sub_phrase_terms)
                elif obj_name in ['names','concepts']:
                    all_terms = [x for x in heading.split(' ') if x]
                    heading = ' & '.join(all_terms)
            if search_phrase:
                for char in '0123456789':
                    if char in search_phrase and int(char) < len(all_terms):
                        search_phrase = search_phrase.replace(char,all_terms[int(char)])
                if '%%' in search_phrase:
                    search_phrase = search_phrase.replace('%%',' & '.join(all_terms))
                heading = search_phrase
                search_phrase = True
            else:
                search_phrase = False
            if LEFT_ITALIC in heading or RIGHT_ITALIC in heading:
                heading = heading.replace(LEFT_ITALIC,'').replace(RIGHT_ITALIC,'')
            return heading,search_phrase
            
        replace_term = '%%%%%%'
        for x in [self.index,
                  self.concepts,
                  self.titles,
                  self.names,
                  self.project_file,
                  self.text_object]:
            if x:
                menu = menu.replace(replace_term,'')
            else:
                menu = menu.replace(replace_term,'*')
            replace_term = replace_term[0:-1]




        new_inp = ''
        # This is the value returned by the function;
        # IF a command, then runs the command at next iteration
        # If FALSE then QUITS
        # If '' then QUERIES
        

        if not inp:
            print(menu)
            inp = numeric_input('?',lower=0,upper=35,alert=False)
        
            

        def complete_file_name (x,first_ending='TXTOB',second_ending='.pkl'):

            """Adds the appropriate suffix to a file name"""
            
            if not x.endswith(first_ending+second_ending):
                if not x.endswith(second_ending):
                    if not x.endswith(first_ending):
                        x += first_ending+second_ending
                    else:
                        x += second_ending
                elif not first_ending in x:
                    x = x.replace(second_ending,
                                  first_ending+second_ending)
            return x

        def show_pages (page_list=None,sentences=False,terms=None,show_all=False,abridge=False,select=False,classifier_phrase=None,show_reversed=False):
            

            """Shows pages in the index"""

            

            if select:
                if classifier_phrase:
                    classifiers = classifier_phrase
                else:
                    classifiers = input('What classifiers will you use?\nCLASSIER1/TERM1 TERM2 etc,CLASSIFER2,...')
                if ';' in classifiers:
                    splitter = ';'
                else:
                    splitter = ','
                classifiers = {x for x in classifiers.split(splitter) if x.strip()}
            else:
                classifiers = set()
                
            return_pages_classified= {}
            
            for c in classifiers:
                return_pages_classified[c] = set()
            
            return_pages = set()
            rejected = set()
            finished_page_stack = Stack()
            classifier_stack = Stack()

            if page_list:
                page_list_saved = list(page_list)
            else:
                page_list_saved = None

            if not sentences:

                all_pages = sort_all_pages(self.text_object.pages.keys())
##                roman_pages = [int_to_roman(x)
##                               for x in sorted([rom_to_int(x) for x in all_pages
##                                                if not x.isnumeric()])]
##                arabic_pages = [str(x) for x in sorted([int(x) for x in all_pages
##                                                        if x.isnumeric()])]
##                all_pages = roman_pages + arabic_pages

            else:

                all_pages = sorted(list(self.text_object.sentence_dict.keys()))
                
            if not page_list:
                page_at = 0
            else:
                page_at = all_pages.index(page_list[0])
            max_len = len(all_pages)

            def convert (x):

                """In x is the index of a sentence,
                it returns the corresponding page index"""

                if not sentences:
                    return x
                return self.text_object.sentence_dict[x][2]

            def get_dehyphenated (terms):

                """Returns words that are found in the text with the hyphen removed"""
                to_add = []
                for x in terms:
                    if x in self.text_object.dehyphenated:
                        to_add.append(self.text_object.dehyphenated[x])
                if isinstance(terms,list):
                    return terms+to_add
                return terms.union(set(to_add))

            def add_classification(xx,page_at):

                """Adds a classification of an index subtype to
                the dictionary keeping track of these"""
                
                if not xx in return_pages_classified:
                                    return_pages_classified[xx]={convert(all_pages[page_at])}
                else:
                    return_pages_classified[xx].add(convert(all_pages[page_at]))
                print('ADDED',xx,convert(all_pages[page_at]))
                classifier_stack.add((xx,convert(all_pages[page_at])))

            def delete_classification(xx,page):

                """Deletes classification"""
                
                if xx:
                    if xx in return_pages_classified:
                        return_pages_classified[xx].discard(page)
                        
                        print('DELETED',xx,page)
                    
                else:
                    return_pages.discard(page)

            
            def classify ():

                """Queries for subclassification of a given index term"""

                all_classes = sorted(set(return_pages_classified.keys()).union(classifiers))
                recommended = recommend_classifiers(show_text,return_pages_classified)
                recommend = ''
                if recommended:
                    print('RECOMMENDED =',','.join(recommended))
                    recommend = sorted(recommended)[0]
                    
                for counter, p in enumerate(all_classes):
                    print(counter,(p in recommended)*'*'+p)
                
                    
                x = input('SPACE to CLASSIFY AS '+recommend+' or ENTER NEW or (A)dd?')
                if x == 'A':
                    return_pages.add(convert(all_pages[page_at]))
                    classifier_stack.add(('',convert(all_pages[page_at])))
                    print(convert(all_pages[page_at]),' ADDED')
                else:
                    if x == ' ':
                        add_classification(recommend,page_at)
                    else:
                        for xx in x.split(','):
                            if xx.isnumeric() and 0<=int(xx)<len(all_classes):
                                xx = all_classes[int(xx)]
                            else:
                                xx = xx.strip()
                            add_classification(xx,page_at)


            def format_indexed_terms (terms,classes=['names','titles','concepts']):

                surround = lambda x:'('+x+')'

                def reform (x,delete=False):

                    """Elimiates bracketed type and search phrase from term"""

                    if delete:
                        dummy,x = get_if(x,'(',')')
                    x= x.split(';;')[0]
                    return x
                
                sorted_terms = {}
                for c in classes:
                    sorted_terms[c] = set()
                    

            
                for x in terms:
                    for c in classes:
                        if surround(c) in x:
                            
                            sorted_terms[c].add(reform(x,delete=True))
                            
                text_list = []
                for c in classes:

                    text_list.append(line_form('<<<'+c+'>>>'+': \n'+'\n'.join(sorted(sorted_terms[c],key=lambda x:x.lower()))))
                    text_list.append('')
                return '\n'.join(text_list)


            def purge_searched_words (words_found,words_searched):
                words_found = set(words_found)
                words_searched = set(words_searched)

                
                def is_found (word):
                    if word in words_found or word.replace('-','') in words_found:
                        return True
                    for w in words_found:
                        if w in word:
                            return True
                    for w in words_searched:
                        if word in w and word.strip()!=w.strip():
                            return True
                    if word in {'titles','words','concepts'}:
                        return True
                    return False
                return {w for w in words_searched if not is_found(w)}
            
                
                    


            page_iterator = None
            
            while all_pages:

                
                
                indexed_terms = {}
                
                if show_reversed and self.reverse_table:
                    if convert(all_pages[page_at]) in self.reverse_table:
                        indexed_terms = self.reverse_table[convert(all_pages[page_at])]
                if not sentences:
                    show_text = self.text_object.get_page(all_pages[page_at])
                else:
                    show_text = self.text_object.get_sentence(all_pages[page_at])

                
                words_from_terms = ' '.join(indexed_terms).replace('.','')
                
                for char in string.punctuation:
                    words_from_terms = words_from_terms.replace(char,' ')
                words_from_terms = [x for x in words_from_terms.split(' ')
                                    if x and x not in SMALL_WORDS and len(x)>4]

                
                if not words_from_terms and terms:
                    words_from_terms = terms
                 

                words_from_terms = get_dehyphenated(words_from_terms)
                
                show_text,found_word_set = bracket(show_text,words_from_terms)
                
                show_text = side_marks (show_text)
                if abridge:
                    show_text = abridge_page(show_text)

                if len(show_text.split('\n'))<ROWS_IN_PAGE:
                    show_text += '\n'*(ROWS_IN_PAGE-len(show_text.split('\n')))
                
               
                
                display = IndexDisplay(columns=2,automatic=False,column_size=[(0,0),(20,40)])
                print(DIVIDER+str(all_pages[page_at])+'/'+convert(all_pages[page_at])+'\n'+DIVIDER)
                display.load(show_text,0)
                
                if self.reverse_table:
                    if convert(all_pages[page_at]) in self.reverse_table:
                        reverse_terms = format_indexed_terms(indexed_terms)
                    display.load(reverse_terms+'\nNOT FOUND\n'+'\n'.join(purge_searched_words(found_word_set,indexed_terms)),1)

                print(display.show())
                
                if not page_list and not page_list_saved:
                    
                    inp = input('<> PAGE # (Q)uit (A)dd (C)lassify OR ENTER A RANGE, OR COMMA TO CLEAR RANGE')
                    if inp:
                        if inp[0].upper() in ['Q','A','C']:
                            inp = inp[0].upper()
                                    
                          
                
                    if not inp:
                        if not page_iterator:
                            page_at += 1
                        else:
                            new_page = next(page_iterator)
                            page_at = all_pages.index(new_page)
                    elif inp == 'Q':
                        break
                    elif ',' in inp or '-' in inp:
                        if inp == ',':
                            page_iterator = None
                        else:
                        
                            page_iterator = itertools.cycle(sort_all_pages(de_range(inp)))
                        
                        
                        
                    elif inp[0] == '<':
                        page_at -= len(inp)
                    elif inp[0] == '>':
                        page_at += len(inp)
                    elif inp in all_pages:
                        page_at = all_pages.index(inp)
                    elif page_at == max_len:
                        page_at = 0
                    elif page_at == -1:
                        page_at = max_len - 1
                    elif inp in ['A',' ']:
                        return_pages.add(convert(all_pages[page_at]))
                        classifier_stack.add(('',convert(all_pages[page_at])))
                        print(convert(all_pages[page_at]),' ADDED')
                    elif inp == 'C':
                        classify()

                                
                    elif not inp in ['A',' ','C']:
                        rejected.add(convert(all_pages[page_at]))
                    
                
                else:
                    if not page_list:
                        if not show_all:
                            page_list = page_list_saved
                        else:
                            break
                    
                    
                    if not show_all or (show_all and select):
                        inp = ''
                        while True:
                            try:
                                inp = input('ANY KEY to ADVANCE or (Q)uit (A)dd (C)lassify (B)ack (U)ndo')
                                if inp == 'U' and classifier_stack.exists():
                                    popped = classifier_stack.pop()
                                    c,p = popped[0],popped[1]
                                    delete_classification(c,p)
                                    
                                                       
                                elif inp:
                                    inp = inp[0].upper()
                                    break
                                else:
                                    break
                            except:
                                print('ERROR')
                        if not inp  in ['A',' ','C','U']:
                            rejected.add(convert(all_pages[page_at]))
                        if inp == 'Q':
                            break
                        elif inp in ['B','<']:
                            if finished_page_stack.exists():
                                popped = finished_page_stack.pop()
                                page_at = all_pages.index(popped)
                                page_list = [popped]+page_list
                                
                        elif inp in ['A',' ']:
                            return_pages.add(convert(all_pages[page_at]))
                            classifier_stack.add(('',convert(all_pages[page_at])))
                            print(convert(all_pages[page_at]),' ADDED')
                        elif inp == 'C':
                            classify()
                        if not page_list:
                            break
                                
                    format_result(search_term='',page_set=return_pages,page_dict=return_pages_classified,rejected=rejected)
                    print('DONE',format_range(set(finished_page_stack.show())))
                    
                    if page_list and not inp in ['B','<','U']:
                        finished_page_stack.add(page_list[0])
                        page_list = page_list[1:]
                        if page_list:
                            page_at = all_pages.index(page_list[0])
            return return_pages, return_pages_classified,rejected 
                    
        #FOR SELECTION THE ACTION CORRESPONDING TO THE INPUT 

        if inp in ['2','3']:

            #Save project, or save project as

            if not self.project_name or inp == '0':
                self.project_name = input('Name of file to save to? ')
                self.project_name = complete_file_name (self.project_name,first_ending='PROB')
                
            if yes_no_input('SAVE TO '+self.project_name):
                if self.project_object:
                    self.project_file = open(directoryname+index_folder+self.project_name,'wb')
                    pickle.dump(self.project_object,
                                self.project_file)
                    self.project_file.close()
        elif inp == '31':
            if yes_no_input('Pre-purge names?'):
                self.text_object.capitalized_phrases = pre_purge(self.text_object.capitalized_phrases)
                    
        elif inp in  ['32','33','34']:
 
            
            while True:
                filename = input('Filename?')
                if filename:
                    break

            obj = {'32':self.project_object.names,
                   '33':self.project_object.titles,
                   '34':self.project_object.concepts}[inp]
            

            all_headers = read_excel(filename=filename)
            scroll = Select()
            all_headers = scroll.scroll_through(all_headers)

            while True:
                temp_inp = input('(A)dd or (R)eplace')
                if temp_inp in ['A','R']:
                    break
            if temp_inp in 'A':

                for x in all_headers:
                    print('ADDING ',x)
                    obj['keep'].add(x)
            else:
                obj['keep'] = set(all_headers)
                             

        elif inp == '1':

            #Open project
            
            self.project_name = input('Name of project to open? ')
            self.project_name = complete_file_name (self.project_name,first_ending='PROB')
            self.project_file = open(directoryname+index_folder+self.project_name,'rb')
            self.project_object = pickle.load(self.project_file)
            self.project_file.close()
            self.searcher_object = Searcher(self.text_object)

        elif inp == '0':

                if not self.text_object or yes_no_input('Overwrite existing TEXT OBJECT?'):

                    if yes_no_input('Load existing TEXT OBJECT '):
                        self.text_filename = complete_file_name(input('Name of project to open?'))
                        tempfile = open(directoryname+index_folder+self.text_filename,'rb')
                        self.text_object = pickle.load(tempfile)
                        tempfile.close()

                    else:
                    
                        self.text_object = Reader()
                        pdf_filename = input('PDF TO READ? ')
                        self.text_object.load(pdf_filename)
                        if not yes_no_input('SAVE as '+complete_file_name(pdf_filename)+'? '):
                            self.text_filename = complete_file_name(input('Save as? '))
                        else:
                            self.text_filename = complete_file_name(pdf_filename)
                        print(index_folder+self.text_filename)
                        tempfile = open(directoryname+index_folder+self.text_filename,'wb')
                        pickle.dump(self.text_object,tempfile)
                        tempfile.close()
                    if not self.project_object:
                        if yes_no_input('Initiate new project? '):
                            self.project_object = Headings(self.text_object)
                            self.searcher_object = Searcher(self.text_object)

                    if not self.index:
                        if yes_no_input('LOAD INDEX? '):
                            inp = input('INDEX NAME? ')
                            self.index_name = complete_file_name (inp,first_ending='IND')
                            tempfile = open(directoryname+index_folder+self.index_name,'rb')
                            self.index = pickle.load(tempfile)
                            tempfile.close()
                            
                            
                    
        elif inp == '4':

            if self.text_object:
                self.project_object = Headings(self.text_object)
                self.searcher_object = Searcher(self.text_object)
            else:
                print('YOU MUST FIRST LOAD TEXT BEFORE YOU CAN INITIATE A PROJECT!')

        elif inp == '5':

            
                    
            if self.project_object:
                self.project_object.define_names()
                self.names = True
                
            else:
                print('NO PROJECT')
        elif inp == '6':
            if self.project_object:
                self.project_object.define_titles()
                self.titles = True 
            else:
                print('NO PROJECT')
        elif inp == '7':
            if self.project_object:
                self.project_object.define_concepts()
                self.concepts = True

        elif inp in ['20','21', '22']:

            print(SEARCH_HELP)

            show_reversed = False
            if self.reverse_table:
                show_reversed = yes_no_input('Show indexed terms on each page?')

            new_inp = run_search (inp,search_term=None,show_reversed=show_reversed)[0]
            

        elif  inp == '35':
            inp = int(numeric_input('Enter search mode.\n1) Pages\n(2) Pages/sentences\n(3) Sentences\n',lower=1,upper=3)) + 19
            inp = str(inp)
            filename = input('FILENAME or names or titles or concepts?')
            if filename not in ['names','titles','concepts']:
                
            
            
                tempfile = open(directoryname+index_folder+filename+'.txt','r',
                                encoding='utf-8')
                file_text = tempfile.read()
                all_terms = reversed(file_text.split('\n'))
            else:
                
                all_terms =  list({'names':self.project_object.names['keep'],
                                   'titles':self.project_object.titles['keep'],
                                   'concepts':self.project_object.concepts['keep']}[filename])
                
                
##            except:
##                print('FILELOAD FAILED')
           
            skip_predetermined = yes_no_input('Skip over if pages have already been identified?')
            add_automatically = skip_predetermined and yes_no_input('Automatically add if pages already identified?')
            
            
            all_results = []
            
            term_stack = Stack()
            for x in all_terms:
                term_stack.add(x)
            

            go_on = True
            while go_on and term_stack.exists():
                x = term_stack.pop()
                go_on, all_results = run_multi_search(x,all_results=all_results,inp=inp)
                
            print('\n'.join(all_results))
            if yes_no_input('Load back into project?'):
                
                if not filename in ['names','titles','concepts']:
                    filename = type_input('Reload into?',
                                          must_be=['names','titles','concepts'],
                                          truncate=False,return_empty=True)
                working_set = {'names':self.project_object.names['keep'],
                               'titles':self.project_object.titles['keep'],
                               'concepts':self.project_object.concepts['keep']}[filename]
                query_when_replacing = yes_no_input('Ask before replacing?')
                
            
                for counter,term in enumerate(all_results):

                    print(str(counter)+':',term)

                    dummy, original = get_if(term,'{','}')
                    print(original)
                    if original.strip() in working_set:
                        
                        working_set.remove(original.strip())
                        working_set.add(term)
                    else:
                        working_set.add(term)
                print(working_set)
                                
                        
                        

                        
            
            
                
          

        elif inp == '8':
            
            object_dict = {'names':self.project_object.names['keep'],
                          'titles':self.project_object.titles['keep'],
                          'concepts':self.project_object.concepts['keep']}
            to_search_object = {'names':self.text_object.words,
                          'titles':self.text_object.title_dict,
                          'concepts':self.text_object.words}
            alternate_object = self.text_object.words

            
            query = yes_no_input('Query?')
            upper_threshold = int(numeric_input(prompt='Last indexed page?'))

            def exclude_pages (page_set,threshold):


                for x in set(page_set):
                    if x.isnumeric() and int(x)>threshold:
                        page_set.remove(x)
                return page_set
            print_during = yes_no_input('Output results while searching?')

            for obj_name in self.index:
                if yes_no_input('\nCreate index for '+obj_name+'? '):
                    searching_set = object_dict[obj_name]
                    searching_object = to_search_object[obj_name]
                    alternative_object = to_search_object['titles']
                    sentence_object = self.text_object.words_in_sentences
                    searching_set = sorted(searching_set)
                    length_of_set = str(len(searching_set))
                    for counter,x in enumerate(searching_set):
                        try:
                        
                            subtract_set_results = False
                            add_set_results = False 

                            if '_' not in x:
                                set_result_phrase, x = get_if(x,left='{',right='}')
                            else:
                                before,after = x.split('_')
                                before_search,before_phrase = get_if (before,left='{',right='}')
                                after_search,after_phrase = get_if (after,left='{',right='}')
                                set_result_phrase = after_search
                                x = before_phrase.strip()+'_'+after_phrase

                            reformed, searchphrase = reform_term(x,obj_name=obj_name)                     

                            
                            if not searchphrase:
                                try:
                                    results = self.searcher_object.search_entry(reformed,searching_object,alternative_object)[0]
                                except:
                                    print(x,'/',reformed,'/','FAILED')
                            elif not reformed.startswith('^'):
                                try:
                                    results = self.searcher_object.search_entry(reformed,searching_object,alternative_object)[0]
                                except:
                                    print(x,'/',reformed,'/','FAILED')
                            else:
                                
                                try:
                                    results = self.text_object.get_pages_for_sentences(self.searcher_object.search_entry(reformed[1:],sentence_object)[0])
                                except:
                                    print(x,'/',reformed,'/','FAILED')
                            if set_result_phrase:
                                operation = '='
                                # + adds pages to existing set, - removes them, = or NOTHING sets to equal,
                                # and % deletes all pages
                                if set_result_phrase[0] in ['+','-','=','%']:
                                    operation = set_result_phrase[0]
                                    set_result_phrase = set_result_phrase[1:]
                                if set_result_phrase:
                                    set_result_pages = de_range(set_result_phrase)
                                if operation == '=':
                                    results = set_result_pages
                                elif operation == '+':
                                    results = results.union(set_result_pages)
                                elif operation == '%':
                                    #to exclude all values
                                    results = set()
                                else:
                                    results -= set_result_pages
                                     
        
                                     
                            if print_during:  
                                print(str(counter)+'/'+length_of_set+' : '+x +'='+reformed)
                                print('    =>'+format_range(results))
                            else:
                                print(str(counter)+'.',end='')
                                

     
                            
                            if query:
                                kept_results = set()
                                for page in results:
                                    print(self.text_object.get_page(page))
                                    print('_______________________________________________')
                                    inp = input('RETURN to KEEP, BLANK+RETURN TO REJECT, Q to KEEP AND STOP QUERYING')
                                    if inp == 'Q':
                                        query = False
                                        kept_results.add(page)
                                    elif not inp:
                                        kept_results.add(page)
                                    if not query:
                                        break
                                results = kept_results

                            #TO RECORD RESULTS 
                            self.index[obj_name][x] = exclude_pages(set(results),threshold=upper_threshold)
                        except:
                            print(counter,x,'FAILED')


                    
                            
                                
        elif inp == '9':

            for counter, object_type in enumerate([self.project_object.names,
                                self.project_object.titles,
                                self.project_object.concepts]):

                heading = {0:'NAMES',
                           1:'TITLES',
                           2:'CONCEPTS'}[counter]

                print(DIVIDER+heading+'\n'+DIVIDER)
                
                for counter2,term in enumerate(object_type['keep']):
                    print(str(counter2)+': '+term)
                
                print(DIVIDER)
                
                   
        elif inp == '12':

            formatted_index = FormatIndex(self.index)
            formatted_index.generate_dictionary()
            if yes_no_input('DO YOU WANT TO CHECK CROSS REFERENCES?'):
                formatted_index.check_cross_references()
            self.index_text = formatted_index.print_dictionary(exclude_empty=yes_no_input('EXCLUDE EMPTY?'),exclude_not_matching=yes_no_input('EXCLUDE UNMATCHED CROSSREFERENCES?'))
            if yes_no_input('DO YOU WANT TO SHOW RESULTS?'):
                print(self.index_text)

            
                
            

        elif inp in ['10','11','26']:


            if inp in ['10','26']:
                file_type = 'keep'
            else:
                file_type = 'purge'
            def display (display_obj,heading='',show_counter=False):
                displaylist = []

                print(heading)
                print()
                for counter, x in enumerate(sorted(display_obj)):

                    print((str(counter)+': ')*show_counter+x)
                    displaylist.append(x)
                print('_______________________________')
                print()
                return displaylist

            if self.project_object:

                if inp == '10':

                    show_counter = yes_no_input('SHOW COUNTER?')
                    

                    display(self.project_object.names[file_type],'NAMES',show_counter=show_counter)
                    display(self.project_object.titles[file_type],'TITLES',show_counter=show_counter)
                    display(self.project_object.concepts[file_type],'CONCEPTS',show_counter=show_counter)
                if inp == '26':
                    while True:

                        sub_inp = type_input(prompt='WHAT FILE TO SAVE? \n (N)ames, (T)titles, (C)concepts, or (A)bort',must_be=['N','T','C'])

                        if not sub_inp == 'A':

                            if sub_inp == 'N':
                                results = display(self.project_object.names[file_type],'NAMES')
                            elif sub_inp == 'T':
                                results = display(self.project_object.titles[file_type],'TITLES')
                            elif sub_inp == 'C':
                                results = display(self.project_object.concepts[file_type],'CONCEPTS')
                            
                            resulttext = '\n'.join(results)
                            print(resulttext)
                            filename = input('FILENAME or RETURN for PROJECTNAME?')
                            if not filename:
                                filename = self.project_name
                            filename += {'N':'_NAMES',
                                         'T':'_TITLES',
                                         'C':'_CONCEPTS'}[sub_inp]
                            tempfile = open(directoryname+index_folder+filename+'.txt','w',
                                            encoding='utf-8')
                            tempfile.write(resulttext)
                            tempfile.close()
  
                    
                        
        elif inp in ['27']:
                sub_inp = type_input(prompt='WHAT FILE TO LOAD? \n (N)ames, (T)titles, (C)concepts, or (A)bort',must_be=['N','T','C','A'])
                if not sub_inp == 'A':
                    heading ={'N':'_NAMES',
                              'T':'_TITLES',
                              'C':'_CONCEPTS'}[sub_inp]
                    working_obj = {'N':self.project_object.names,
                              'T':self.project_object.titles,
                              'C':self.project_object.concepts}[sub_inp]
                    filename = input('FILENAME or RETURN for PROJECTNAME?')
                    if not filename:
                        filename = self.project_name
                    
                    tempfile = open(directoryname+index_folder+filename+heading+'.txt','r',
                                    encoding='utf-8')
                    file_text = tempfile.read()
                    file_set = file_text.split('\n')

                    for x in working_obj['keep']:
                        if x not in file_set: 
                            working_obj['purge'].add(x)
                    working_obj['keep'] = set(file_set)
                    for x in file_set:
                        if x in working_obj['purge']:
                            working_obj['purge'].remove(x)


        elif inp in ['13']:

            show_reversed = False
            
            if self.reverse_table:
                show_reversed = yes_no_input('Show indexed terms on each page?')
            show_pages(show_reversed=show_reversed)

            
        elif inp in ['14']:

            self.reverse_index()

        elif inp in ['15']:

                 

            classify_terms = yes_no_input('Classify indexed terms?')

            def print_reverse_index (classify=classify_terms):

                """Prints out a reverse index"""

                surround = lambda x:'('+x+')'
                

                def reform (x,delete=False):

                    """Elimiates bracketed type and search phrase from term"""

                    if delete:
                        dummy,x = get_if(x,'(',')')
                    x= x.split(';;')[0]
                    return x
                

                def format_reversed_terms (terms,classes=['names','titles','concepts'],classify=classify_terms,delete=True):

                    """Formats the reversed index entry for a given page"""

                    if classify:
                        
                        sorted_terms = {}
                        for c in classes:
                            sorted_terms[c] = set()
                            

                    
                        for x in terms:
                            for c in classes:
                                if surround(c) in x:
                                    
                                    sorted_terms[c].add(reform(x,delete=delete))
                                    
                        text_list = []
                        for c in classes:

                            text_list.append(line_form('<<<'+c+'>>>'+': \n'+'\n'.join(sorted(sorted_terms[c],key=lambda x:x.lower()))))
                            text_list.append('')
                        return '\n'.join(text_list)

                    else:

                        return line_form(';'.join(sorted([reform(x,delete) for x in terms])))
                        
               
                text_list = []
                for page in sort_all_pages(self.reverse_table.keys()):
                    

                    text_list.append('_'*80)
                    text_list.append('PAGE ='+page)
                    if page in self.reverse_table:

                        text_list.append(format_reversed_terms(self.reverse_table[page],classify=classify_terms))
                    else:
                        print(page,'NOT FOUND')
                    
                return text_list
            self.reverse_index_text = '\n'.join(print_reverse_index(classify_terms))
            print(self.reverse_index_text)
            
        
        elif inp in ['16']:

            for index_type in self.index:

                print('REVIEWING '+index_type+'\n'+DIVIDER)
                for entry in self.index[index_type]:
                    inp = type_input(prompt='(R)EVIEW, (Q)UIT, OR (D)ELETE '+entry,must_be=['R','Q','R'])
                    if inp == 'D':
                        del self.index[index_type][entry]
                        
                    elif inp == 'R':
                        for page in sorted((self.index[index_type][entry])):
                            
                            show_text = self.text_object.get_page(page)
                            words_from_terms = entry
                            for char in string.punctuation:
                                words_from_terms = words_from_terms.replace(char,' ')
                            words_from_terms = [x for x in words_from_terms.split(' ') if x and not x in SMALL_WORDS]
                            words_from_terms = get_dehyphenated(words_from_terms)
                            show_text, found_word_set = bracket(show_text,words_from_terms)
                            show_text = side_marks(show_text)
                            
                            print(DIVIDER+'\n'
                                  +'[['+page+']]'+'\n'
                                  +DIVIDER+'\n'
                                  +show_text+DIVIDER)
                            if yes_no_input('SPACE OR Y(ES) to REJECT?'):
                                self.index[index_type][entry].remove(page)
                        inp = input('PAGES TO ADD?')
                        if inp:
                            pages = inp.split(',')
                            for p in pages:
                                if p in self.text_object.pages.keys():
                                    print('Adding ',p)
                                    self.index[index_type][entry].add(p)
                    elif inp == 'Q':
                        break 
                        
        elif inp in ['17']:
            self.index_name = complete_file_name (self.project_name.replace('PROB','').replace('.pkl',''),first_ending='IND')
            name_inp = input('RETURN TO USE '+self.index_name+' AS FILENAME, OR ENTER NEW NAME? ')
            if len(name_inp) > 3:
                self.index_name =  complete_file_name (name_inp,first_ending='IND')
            
            if self.index_name and (input('SAVE TO '+self.index_name)in ['yes','y','Y','Yes',' ']):
                if self.project_object:
                    tempfile = open(directoryname+index_folder+self.index_name,'wb')
                    pickle.dump(self.index,tempfile)
                    tempfile.close()

        elif inp in ['18']:

            if yes_no_input('LOAD INDEX? '):
                        inp = input('INDEX NAME? ')
                        self.index_name = complete_file_name (inp,first_ending='IND')
                        tempfile = open(directoryname+index_folder+self.index_name,'rb')
                        self.index = pickle.load(tempfile)
                        tempfile.close()

        elif inp in ['19']:

            def save_text_file (x,first_ending='FIN'):

                """To save given text file"""

                
                filename = complete_file_name(self.project_name.replace('PROB','').replace('.pkl',''),
                                              first_ending=first_ending,
                                              second_ending='.txt')
                while True:
                    if yes_no_input('SAVE to '+filename+'? '):
                        break
                    else:
                        filename = complete_file_name(input('Filename? '),
                                                      first_ending=first_ending,
                                                      second_ending='.txt')
                tempfile = open(directoryname+index_folder+filename,'w',
                                encoding='utf-8')
                tempfile.write(x)
                tempfile.close()

            t_inp = type_input('(I)ndex or (R)everse index',must_be=['I','R'],return_empty=True)
            if t_inp == 'I':
                if self.index_text:
                    if yes_no_input('SAVE FORMATTED INDEX as TEXTFILE'):
                        save_text_file (self.index_text)
                        print('SAVED')
                else:
                    print('INDEX NOT YET FORMATTED!')
            elif t_inp == 'R' and self.reverse_index_text:
                if self.reverse_index_text:
                    if yes_no_input('SAVE RETURN INDEX as TEXTFULE'):
                        save_text_file (self.reverse_index_text,first_ending='REV')
                else:
                    print('REVERSE INDEX NOT YET FORMATTED!')
                    
               

        

                


##        elif inp in ['25']:
##            if input('QUIT? ') in YESTERMS:
##                return False
        elif inp in ['23']:
            self.override = not self.override
            print('OVERRIDE ',{True:'ON',
                   False:'OFF'}[self.override])

        elif inp in ['28']:
            prompt = """
TO CHANGE STATUS 

(n)ames
(i)italicized phrases
(q)uoted phrases
(p)arenthetical phrases 1
p(a)renthetical phrases 2"""
            
            while True:
                obj_type = input(prompt)

                if obj_type == 'n':
                    print('NAMES')
                    self.project_object.names_done = not self.project_object.names_done
                elif obj_type == 'i':
                    self.project_object.italicized_phrases_done = not self.project_object.italicized_phrases_done
                elif obj_type == 'q':
                    self.project_object.quoted_phrases_done = not self.project_object.quoted_phrases_done
                elif obj_type == 'p':
                    self.project_object.paren_phrases1_done = not self.project_object.paren_phrases1_done
                elif obj_type == 'a':
                    self.project_object.paren_phrases2_done = not self.project_object.paren_phrases2_done
                else:
                    break
        elif inp in ['29']:

            for x in sorted(self.text_object.dehyphenated):
                print (x,'=',self.text_object.dehyphenated[x])
        elif inp in ['25']:
            self.reverse_index()
            if yes_no_input('Pages as indexes? '):
                page_as_index = True
            else:
                page_as_index = False
                
            def convert (x):

                if x.isnumeric():
                    return int(x)
                else:
                    return rom_to_int(x)-1000
                
            def uncovert (x):
                if x>=0:
                    return str(x)
                else:
                    return int_to_roman(x+1000)
                
                

            counter = 1
            for counter, page in enumerate(sorted([convert(x) for x in self.reverse_table.keys()])):
                page = uncovert(page)
                page_keys = self.reverse_table[page]
                if page.isnumeric():
                    if page_as_index:
                        page_index = 'b.'+str(page)
                    else:
                        page_keys.add('page@'+page)
                        page_index = 'c.'+str(counter)
                else:
                    if page_as_index:
                        page_index = 'a.'+str(rom_to_int(page))
                        
                    else:
                        page_index = 'c.'+str(counter)
                        page_keys.add('rpage@'+page)
                page_text = self.text_object.get_page(page).replace('.\n','%%%').replace('\n',' ').replace('%%%','\n')
                
                
                
                self.return_dict[page_index]={'keys':page_keys,
                                         'text':page_text}
            
            if yes_no_input('Quit and return index? '):
                return False
            
            
            
        return new_inp 
            
    def console (self):

        """To call up the console and initiate main loop"""

        inp = ''
        print(READ_ME)
        while True:

            if self.override:
                try:
                    inp = self.input_term(inp)
                    if inp is False:
                        break
                    
                except:
                    print('ERROR')
            else:
                inp = self.input_term(inp)
                if inp is False:
                        break
        if self.return_dict and yes_no_input('Return index? '):
            return self.return_dict
        return {}


            
     

if __name__ == "__main__":

    my_indexes = Index_Maker()
    my_indexes.console()


        
    
            
        

        
        
